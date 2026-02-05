"""
Sistema Universal de Exportação - MegaCLI v6.0

Sistema genérico para exportar dados para Excel ou arquivos TXT
baseado em configuração JSON.

Autor: MegaCLI Team
Data: 23/01/2026
Versão: 2.0.0
"""

import pandas as pd
import json
from pathlib import Path
from typing import Dict, List, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


class SistemaExportacao:
    """Sistema universal de exportação baseado em JSON."""
    
    def __init__(self):
        """Inicializa sistema de exportação."""
        self.cores_padrao = {
            'header': 'FFFFFF',
            'header_bg': '4472C4',
            'alta': 'C6EFCE',
            'media': 'FFEB9C',
            'baixa': 'FFC7CE'
        }
    
    def exportar(self, config: Dict[str, Any]) -> bool:
        """
        Exporta dados baseado em configuração JSON.
        
        Args:
            config: Dicionário com configuração completa
            
        Returns:
            True se sucesso, False caso contrário
            
        Exemplo de config:
        {
            "tipo": "excel",  # ou "txt"
            "arquivo": "caminho/arquivo.xlsx",
            "aba": "NOME_ABA",
            "dados": [...],  # Lista de dicionários
            "colunas": ["Col1", "Col2", ...],  # Opcional
            "formatacao": {
                "titulo": "Título da Aba",
                "cores_condicionais": {
                    "coluna": "Score",
                    "regras": [
                        {"condicao": ">=", "valor": 80, "cor": "alta"},
                        {"condicao": ">=", "valor": 60, "cor": "media"},
                        {"condicao": "<", "valor": 60, "cor": "baixa"}
                    ]
                },
                "estatisticas": ["Total", "Média", "Máximo"]
            }
        }
        """
        try:
            tipo = config.get('tipo', 'excel').lower()
            
            if tipo == 'excel':
                return self._exportar_excel(config)
            elif tipo == 'txt':
                return self._exportar_txt(config)
            else:
                print(f"❌ Tipo de exportação inválido: {tipo}")
                return False
                
        except Exception as e:
            print(f"❌ Erro ao exportar: {e}")
            return False
    
    def _exportar_excel(self, config: Dict[str, Any]) -> bool:
        """Exporta para Excel."""
        arquivo = Path(config['arquivo'])
        aba = config['aba']
        dados = config['dados']
        
        # Criar DataFrame
        if 'colunas' in config:
            df = pd.DataFrame(dados, columns=config['colunas'])
        else:
            df = pd.DataFrame(dados)
        
        # Criar arquivo se não existir
        if not arquivo.exists():
            df.to_excel(str(arquivo), sheet_name=aba, index=False)
            wb = load_workbook(str(arquivo))
        else:
            wb = load_workbook(str(arquivo))
            
            # Remover aba se existir
            if aba in wb.sheetnames:
                del wb[aba]
            
            # Criar nova aba
            ws = wb.create_sheet(aba)
        
        ws = wb[aba] if aba in wb.sheetnames else wb.create_sheet(aba)
        
        # Aplicar formatação se especificada
        if 'formatacao' in config:
            self._aplicar_formatacao_excel(ws, df, config['formatacao'])
        else:
            # Formatação básica
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(str(arquivo))
        print(f"✅ Dados exportados para Excel: {arquivo} (aba: {aba})")
        return True
    
    def _aplicar_formatacao_excel(
        self,
        ws,
        df: pd.DataFrame,
        formatacao: Dict[str, Any]
    ) -> None:
        """Aplica formatação ao Excel."""
        row_atual = 1
        
        # Título
        if 'titulo' in formatacao:
            ws[f'A{row_atual}'] = formatacao['titulo']
            ws[f'A{row_atual}'].font = Font(bold=True, size=14, color=self.cores_padrao['header'])
            ws[f'A{row_atual}'].fill = PatternFill(
                start_color=self.cores_padrao['header_bg'],
                end_color=self.cores_padrao['header_bg'],
                fill_type='solid'
            )
            ws.merge_cells(f'A{row_atual}:{chr(64 + len(df.columns))}{row_atual}')
            row_atual += 1
        
        # Estatísticas
        if 'estatisticas' in formatacao:
            for stat in formatacao['estatisticas']:
                if stat == 'Total':
                    ws[f'A{row_atual}'] = f'Total de Registros: {len(df)}'
                elif stat == 'Média' and len(df.select_dtypes(include=['number']).columns) > 0:
                    col = df.select_dtypes(include=['number']).columns[0]
                    ws[f'A{row_atual}'] = f'Média {col}: {df[col].mean():.2f}'
                elif stat == 'Máximo' and len(df.select_dtypes(include=['number']).columns) > 0:
                    col = df.select_dtypes(include=['number']).columns[0]
                    ws[f'A{row_atual}'] = f'Máximo {col}: {df[col].max():.2f}'
                
                ws[f'A{row_atual}'].font = Font(bold=True)
                row_atual += 1
            
            row_atual += 1
        
        # Dados
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), row_atual):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Cabeçalho
                if r_idx == row_atual:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
        
        # Cores condicionais
        if 'cores_condicionais' in formatacao:
            self._aplicar_cores_condicionais(ws, df, formatacao['cores_condicionais'], row_atual)
        
        # Auto-ajustar colunas
        for column in ws.columns:
            max_length = 0
            # Pegar a letra da coluna da primeira célula (que pode ser MergedCell, então precisa de cuidado)
            try:
                # Tenta pegar column_letter da primeira célula
                column_letter = column[0].column_letter
            except AttributeError:
                # Se for MergedCell, tenta obter da coordenada
                import openpyxl.utils
                column_letter = openpyxl.utils.get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _aplicar_cores_condicionais(
        self,
        ws,
        df: pd.DataFrame,
        config_cores: Dict[str, Any],
        row_inicio: int
    ) -> None:
        """Aplica cores condicionais."""
        coluna = config_cores['coluna']
        regras = config_cores['regras']
        
        # Encontrar índice da coluna
        col_idx = list(df.columns).index(coluna) + 1
        
        # Aplicar regras
        for r_idx in range(row_inicio + 1, row_inicio + len(df) + 1):
            valor = ws.cell(row=r_idx, column=col_idx).value
            
            if valor is not None and isinstance(valor, (int, float)):
                cor = None
                for regra in regras:
                    condicao = regra['condicao']
                    valor_ref = regra['valor']
                    
                    if condicao == '>=' and valor >= valor_ref:
                        cor = self.cores_padrao.get(regra['cor'], 'FFFFFF')
                        break
                    elif condicao == '<' and valor < valor_ref:
                        cor = self.cores_padrao.get(regra['cor'], 'FFFFFF')
                        break
                
                if cor:
                    # Aplicar cor em toda a linha
                    for c in range(1, len(df.columns) + 1):
                        ws.cell(row=r_idx, column=c).fill = PatternFill(
                            start_color=cor, end_color=cor, fill_type='solid'
                        )
    
    def _exportar_txt(self, config: Dict[str, Any]) -> bool:
        """Exporta para TXT."""
        arquivo = Path(config['arquivo'])
        dados = config['dados']
        
        with open(arquivo, 'w', encoding='utf-8') as f:
            # Título
            if 'formatacao' in config and 'titulo' in config['formatacao']:
                f.write("=" * 70 + "\n")
                f.write(config['formatacao']['titulo'] + "\n")
                f.write("=" * 70 + "\n\n")
            
            # Dados
            formato_linha = config.get('formato_linha', None)
            for item in dados:
                if formato_linha:
                    # Usar template de formatação
                    try:
                        linha = formato_linha.format(**item)
                    except KeyError:
                        linha = str(item)
                else:
                    linha = str(item)
                f.write(linha + "\n")
            
            # Rodapé
            f.write("\n" + "=" * 70 + "\n")
            f.write(f"Total: {len(dados)} registros\n")
            f.write("=" * 70 + "\n")
        
        print(f"✅ Dados exportados para TXT: {arquivo}")
        return True


# Função helper para criar configuração de jogos
def criar_config_jogos(
    arquivo: str,
    jogos: List[Dict],
    aba: Optional[str] = None
) -> Dict[str, Any]:
    """
    Cria configuração JSON para exportar jogos.
    
    Args:
        arquivo: Caminho do arquivo Excel
        jogos: Lista de jogos
        aba: Nome da aba (opcional, usa timestamp se None)
        
    Returns:
        Dicionário de configuração
    """
    if aba is None:
        aba = f"JOGOS_{datetime.now().strftime('%Y%m%d')}"
    
    # Converter jogos para formato tabular
    dados = []
    for jogo in jogos:
        dados.append({
            'Rank': jogo['rank'],
            'Num1': jogo['numeros'][0],
            'Num2': jogo['numeros'][1],
            'Num3': jogo['numeros'][2],
            'Num4': jogo['numeros'][3],
            'Num5': jogo['numeros'][4],
            'Num6': jogo['numeros'][5],
            'Score': jogo['score'],
            'Probabilidade': jogo['probabilidade'],
            'Confiança': jogo['confianca'].split()[0]
        })
    
    return {
        "tipo": "excel",
        "arquivo": arquivo,
        "aba": aba,
        "dados": dados,
        "formatacao": {
            "titulo": f"JOGOS GERADOS - {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            "estatisticas": ["Total", "Média"],
            "cores_condicionais": {
                "coluna": "Score",
                "regras": [
                    {"condicao": ">=", "valor": 80, "cor": "alta"},
                    {"condicao": ">=", "valor": 60, "cor": "media"},
                    {"condicao": "<", "valor": 60, "cor": "baixa"}
                ]
            }
        }
    }


# Exports
__all__ = ['SistemaExportacao', 'criar_config_jogos']


# Teste standalone
if __name__ == "__main__":
    print("\n🧪 Testando Sistema de Exportação...\n")
    
    sistema = SistemaExportacao()
    
    # Teste 1: Exportar jogos
    jogos_teste = [
        {
            'rank': i,
            'numeros': [5, 10, 15, 20, 25, 30],
            'score': 85.0 - i * 5,
            'probabilidade': 85.0 - i * 5,
            'confianca': 'ALTA ⭐⭐⭐' if i <= 3 else 'MÉDIA ⭐⭐'
        }
        for i in range(1, 11)
    ]
    
    config = criar_config_jogos("test_sistema.xlsx", jogos_teste)
    sistema.exportar(config)
    
    print("\n✅ Teste concluído!\n")
