"""
Exportador Excel Centralizado - MegaCLI v6.0

Módulo único para gerenciar TODAS as gravações no Excel.
Centraliza formatação, estilos e operações de escrita.

Autor: MegaCLI Team
Data: 23/01/2026
Versão: 1.0.0
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


class ExportadorExcel:
    """Gerenciador centralizado de exportações para Excel."""
    
    # Cores padrão
    COR_HEADER = 'FFFFFF'
    COR_HEADER_BG = '4472C4'
    COR_ALTA = 'C6EFCE'  # Verde claro
    COR_MEDIA = 'FFEB9C'  # Amarelo claro
    COR_BAIXA = 'FFC7CE'  # Vermelho claro
    
    def __init__(self, arquivo_excel: str):
        """
        Inicializa exportador com arquivo Excel.
        
        Args:
            arquivo_excel: Caminho do arquivo Excel
        """
        self.arquivo_excel = Path(arquivo_excel)
        
        # Criar arquivo se não existir
        if not self.arquivo_excel.exists():
            wb = pd.ExcelWriter(str(self.arquivo_excel), engine='openpyxl')
            wb.close()
    
    def criar_ou_atualizar_aba(
        self,
        nome_aba: str,
        df: pd.DataFrame,
        substituir: bool = True
    ) -> None:
        """
        Cria ou atualiza uma aba com DataFrame.
        
        Args:
            nome_aba: Nome da aba
            df: DataFrame a gravar
            substituir: Se True, substitui aba existente
        """
        wb = load_workbook(str(self.arquivo_excel))
        
        # Remover aba se existir e substituir=True
        if nome_aba in wb.sheetnames and substituir:
            del wb[nome_aba]
        
        # Criar nova aba
        ws = wb.create_sheet(nome_aba)
        
        # Escrever DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(str(self.arquivo_excel))
    
    def gravar_jogos(
        self,
        jogos: List[Dict],
        nome_aba: str = "JOGOS_GERADOS"
    ) -> None:
        """
        Grava jogos gerados em aba específica com formatação.
        
        Args:
            jogos: Lista de dicionários com jogos
            nome_aba: Nome da aba
        """
        # Criar DataFrame
        dados = []
        for jogo in jogos:
            dados.append({
                'Rank': jogo['rank'],
                'Num1': jogo['numeros'][0],
                'Num2': jogo['numeros'][1],
                'Num3': jogo['numeros'][2],
                'Num4': jogo['numeros'][3],
                'Num5': jogo['numeros'][4],
                'Num6': jogo['numeros'][5],
                'Score': jogo['score'],
                'Probabilidade': jogo['probabilidade'],
                'Confiança': jogo['confianca'].split()[0]  # Remover estrelas
            })
        
        df = pd.DataFrame(dados)
        
        # Gravar
        wb = load_workbook(str(self.arquivo_excel))
        
        # Remover aba se existir
        if nome_aba in wb.sheetnames:
            del wb[nome_aba]
        
        # Criar nova aba
        ws = wb.create_sheet(nome_aba)
        
        # Cabeçalho
        ws['A1'] = f'JOGOS GERADOS - {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        ws['A1'].font = Font(bold=True, size=14, color=self.COR_HEADER)
        ws['A1'].fill = PatternFill(start_color=self.COR_HEADER_BG, end_color=self.COR_HEADER_BG, fill_type='solid')
        ws.merge_cells('A1:J1')
        
        # Estatísticas
        row_atual = 2
        ws[f'A{row_atual}'] = f'Total de Jogos: {len(jogos)}'
        ws[f'A{row_atual}'].font = Font(bold=True)
        row_atual += 1
        
        ws[f'A{row_atual}'] = f'Score Médio: {df["Score"].mean():.2f}'
        row_atual += 2
        
        # Dados
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), row_atual):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                
                # Formatar cabeçalho
                if r_idx == row_atual:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
                
                # Aplicar cores por confiança
                elif c_idx == 10:  # Coluna Confiança
                    score = ws.cell(row=r_idx, column=8).value  # Score
                    if score and isinstance(score, (int, float)):
                        if score >= 80:
                            cor = self.COR_ALTA
                        elif score >= 60:
                            cor = self.COR_MEDIA
                        else:
                            cor = self.COR_BAIXA
                        
                        # Aplicar cor em toda a linha
                        for col in range(1, 11):
                            ws.cell(row=r_idx, column=col).fill = PatternFill(
                                start_color=cor, end_color=cor, fill_type='solid'
                            )
        
        # Auto-ajustar colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(str(self.arquivo_excel))
        print(f"✅ Jogos gravados na aba '{nome_aba}' do Excel")
    
    def gravar_validacao(
        self,
        df_validacao: pd.DataFrame,
        nome_aba: str = "VALIDAÇÃO_100"
    ) -> None:
        """
        Grava validação histórica.
        
        Args:
            df_validacao: DataFrame com validação
            nome_aba: Nome da aba
        """
        self.criar_ou_atualizar_aba(nome_aba, df_validacao)
        
        # Aplicar formatação
        wb = load_workbook(str(self.arquivo_excel))
        ws = wb[nome_aba]
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        wb.save(str(self.arquivo_excel))
        print(f"✅ Validação gravada na aba '{nome_aba}' do Excel")
    
    def gravar_universo(
        self,
        numeros: List[int],
        nome_aba: str,
        metricas: Dict,
        df_historico: pd.DataFrame,
        janela: int = 100
    ) -> None:
        """
        Grava universo reduzido com validação.
        
        Args:
            numeros: Lista de números do universo
            nome_aba: Nome da aba
            metricas: Dicionário com métricas
            df_historico: DataFrame com histórico
            janela: Janela de validação
        """
        # Usar atualizador existente
        from src.validacao.atualizador_abas_universos import criar_aba_universo
        
        criar_aba_universo(
            str(self.arquivo_excel),
            nome_aba,
            numeros,
            df_historico,
            janela,
            verbose=False
        )
        
        print(f"✅ Universo gravado na aba '{nome_aba}' do Excel")


# Exports
__all__ = ['ExportadorExcel']


# Teste standalone
if __name__ == "__main__":
    print("\n🧪 Testando Exportador Excel...\n")
    
    # Criar exportador de teste
    exportador = ExportadorExcel("test_exportador.xlsx")
    
    # Teste 1: Gravar jogos
    jogos_teste = [
        {
            'rank': i,
            'numeros': [5, 10, 15, 20, 25, 30],
            'score': 85.0 - i,
            'probabilidade': 85.0 - i,
            'confianca': 'ALTA ⭐⭐⭐'
        }
        for i in range(1, 11)
    ]
    
    exportador.gravar_jogos(jogos_teste, "TESTE_JOGOS")
    print("✅ Teste de gravação de jogos concluído")
    
    # Teste 2: Gravar validação
    df_teste = pd.DataFrame({
        'Concurso': range(1, 6),
        'Acertos': [3, 4, 5, 3, 4]
    })
    
    exportador.gravar_validacao(df_teste, "TESTE_VALIDACAO")
    print("✅ Teste de gravação de validação concluído")
    
    print("\n✅ Todos os testes passaram!\n")
