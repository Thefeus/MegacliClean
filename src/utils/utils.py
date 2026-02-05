from openpyxl import load_workbook, Workbook
from typing import Dict, Any
import pandas as pd
import logging
from .app_config import config as app_settings

def get_digital_root(n: int) -> int:
    while n >= 10: n = sum(int(digit) for digit in str(n))
    return n

def is_prime(n: int) -> bool:
    if n <= 1: return False
    if n <= 3: return True
    if n % 2 == 0 or n % 3 == 0: return False
    i = 5
    while i * i <= n:
        if n % i == 0 or n % (i + 2) == 0: return False
        i += 6
    return True

def load_params(file_path: str) -> Dict[str, Any]:
    # Estrutura de dados para armazenar valor e origem
    params = {
        'num_backtest_games': {'value': 10, 'source': 'Padrão'},
        'ml_window_size': {'value': 200, 'source': 'Padrão'},
        'elite_feature_count': {'value': 150, 'source': 'Padrão'},
        'somente_ia': {'value': 'N', 'source': 'Padrão'},
        'use_ollama_ia': {'value': 'N', 'source': 'Padrão'},
        'ollama_model': {'value': 'llama3', 'source': 'Padrão'},
        'use_ai_feature_generation': {'value': 'N', 'source': 'Padrão'},
        'run_ollama_direct_prediction': {'value': 'N', 'source': 'Padrão'},
        'ultimo_jogo_analisado': {'value': 0, 'source': 'Padrão'},
        'run_formula_suggestion': {'value': 'N', 'source': 'Padrão'} # Novo parâmetro de análise
    }
    try:
        workbook = load_workbook(file_path)
        sheet_name = 'Param'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            def get_cell_value(row, col, default_val, type_converter):
                cell = sheet.cell(row=row, column=col)
                source = f"{sheet_name}!{cell.coordinate}"
                if cell.value is not None:
                    try:
                        return {'value': type_converter(cell.value), 'source': source}
                    except (ValueError, TypeError):
                        return {'value': default_val, 'source': source + " (valor inválido, usando padrão)"}
                return {'value': default_val, 'source': 'Padrão'}

            params['num_backtest_games'] = get_cell_value(3, 2, params['num_backtest_games']['value'], int)
            params['ml_window_size'] = get_cell_value(4, 2, params['ml_window_size']['value'], int)
            params['elite_feature_count'] = get_cell_value(5, 2, params['elite_feature_count']['value'], int)
            params['somente_ia'] = get_cell_value(7, 2, params['somente_ia']['value'], str)
            params['use_ollama_ia'] = get_cell_value(8, 2, params['use_ollama_ia']['value'], str)
            params['ollama_model'] = get_cell_value(9, 2, params['ollama_model']['value'], str)
            params['use_ai_feature_generation'] = get_cell_value(10, 2, params['use_ai_feature_generation']['value'], str)
            params['run_ollama_direct_prediction'] = get_cell_value(11, 2, params['run_ollama_direct_prediction']['value'], str)
            params['ultimo_jogo_analisado'] = get_cell_value(12, 2, params['ultimo_jogo_analisado']['value'], int)
            params['run_formula_suggestion'] = get_cell_value(13, 2, params['run_formula_suggestion']['value'], str)

        else: # Se a aba "Param" não existe, cria com valores padrão
            param_sheet = workbook.create_sheet(sheet_name)
            headers = ["Parâmetro", "Valor", "Descrição"]
            param_sheet.append(headers)
            
            param_data = [
                ("--- Backtesting ---", "", ""),
                ("Qtd. Jogos p/ Backtest", params['num_backtest_games']['value'], "Número de sorteios recentes para incluir no processo de backtesting."),
                ("Janela de Aprendizado", params['ml_window_size']['value'], "Quantos sorteios passados serão usados para treinar o modelo a cada passo."),
                ("Qtd. Features de Elite", params['elite_feature_count']['value'], "Número de features mais importantes a serem usadas no modelo 'refinado'."),
                ("--- Execução com IA ---", "", ""),
                ("Somente Consultar IA Gemini", params['somente_ia']['value'], "S/N - Pula o backtesting e usa o ML_NEURO existente."),
                ("Usar IA Local Ollama", params['use_ollama_ia']['value'], "S/N - Ativa a geração de apostas com um modelo rodando localmente via Ollama."),
                ("Modelo Ollama", params['ollama_model']['value'], "Nome do modelo Ollama a ser usado (ex: llama3)."),
                ("Gerar Features com IA", params['use_ai_feature_generation']['value'], "S/N - (Experimental) Deixa a IA Ollama gerar features. Lento."),
                ("Executar Predição Direta Ollama", params['run_ollama_direct_prediction']['value'], "S/N - Ignora o pipeline de ML e pede para a IA prever diretamente."),
                ("--- Análise e Refinamento ---", "", ""),
                ("Último Jogo Analisado", params['ultimo_jogo_analisado']['value'], "Checkpoint do último concurso processado pelo fluxo de refinamento."),
                ("Sugerir Fórmula de Feature (IA)", params['run_formula_suggestion']['value'], "S/N - Usa a IA para analisar os neurônios e sugerir uma nova fórmula de feature.")
            ]
            for row_data in param_data:
                param_sheet.append(row_data)

            workbook.save(file_path)
        return params
    except Exception as e:
        logging.error(f"Erro ao carregar parâmetros: {e}", exc_info=True)
        return params

def add_static_indicators(df: pd.DataFrame) -> pd.DataFrame:
    logging.info("Calculando indicadores estáticos...")
    balls = app_settings.get_ball_columns()
    df_indicators = df.copy()
    df_indicators['Soma'] = df_indicators[balls].sum(axis=1)
    df_indicators['Pares'] = df_indicators[balls].apply(lambda r: sum(1 for x in r if x % 2 == 0), axis=1)
    df_indicators['Impares'] = 6 - df_indicators['Pares']
    df_indicators['Primos'] = df_indicators[balls].apply(lambda r: sum(1 for x in r if is_prime(int(x))), axis=1)
    df_indicators['Tesla_Div_3'] = df_indicators[balls].apply(lambda r: sum(1 for x in r if x % 3 == 0), axis=1)
    df_indicators['Tesla_DR_369'] = df_indicators[balls].apply(lambda r: sum(1 for x in r if get_digital_root(int(x)) in [3, 6, 9]), axis=1)
    df_indicators['Tesla_Soma_DR'] = df_indicators['Soma'].apply(lambda x: get_digital_root(int(x)))
    return df_indicators

def setup_refinement_sheet(file_path: str):
    """
    Garante que a aba 'ANALISE_REFINAMENTO' exista no arquivo Excel.
    Se não existir, cria a aba com os cabeçalhos necessários.
    """
    sheet_name = "ANALISE_REFINAMENTO"
    headers = [
        "Concurso", "Data_Predicao", 
        "Predicao_Gemini", "Estrategia_Gemini", "Acertos_Gemini",
        "Predicao_Ollama", "Estrategia_Ollama", "Acertos_Ollama",
        "Resultado_Real", "Refinamento_IA"
    ]

    try:
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            print(f"Criando nova aba '{sheet_name}'...")
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(headers)
            workbook.save(file_path)
    except FileNotFoundError:
        print(f"Arquivo {file_path} não encontrado. Uma nova planilha será criada com a aba de refinamento.")
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet.append(headers)
        # Remove a aba 'Sheet' padrão que é criada automaticamente
        if 'Sheet' in workbook.sheetnames and len(workbook.sheetnames) > 1:
            del workbook['Sheet']
        workbook.save(file_path)
    except Exception as e:
        logging.error(f"Erro ao configurar a aba de refinamento: {e}", exc_info=True)
        print(f"Ocorreu um erro ao configurar a aba '{sheet_name}': {e}")
