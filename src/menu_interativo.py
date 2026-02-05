"""
Menu Interativo MegaCLI v6.3.0

Interface CLI interativa para acesso às funcionalidades do sistema.

Autor: MegaCLI Team
Data: 04/02/2026
Versão: 6.3.0
"""

import sys
from pathlib import Path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT))
sys.path.insert(0, str(PROJECT_ROOT / "src"))

import pandas as pd
from datetime import datetime
from typing import Optional
import os

# Imports do sistema

# Imports do Mapa de Fontes (Desacoplado)
from config.fontes import (
    CONFIG,
    ANALISE_PARAMS,
    GERADOR_JOGOS,
    VALIDADOR_1000_JOGOS,
    PREVISAO_30N,
    CICLO_REFINAMENTO,
    MODO_CONSERVADOR,
    VISUALIZACAO,
    RANKING,
    ANALISADOR_HISTORICO,
    ANALISADOR_9_NUMEROS,
    INDICADOR_OTIMIZADO_10N,
    ANALISADOR_UNIVERSO_REDUZIDO,
    SISTEMA_EXPORTACAO,
    ANALISE_V6
)

# Aliases para manter compatibilidade com código existente e evitar refatoração massiva
ARQUIVO_HISTORICO = CONFIG.ARQUIVO_HISTORICO
RESULTADO_DIR = CONFIG.RESULTADO_DIR
AnaliseConfig = ANALISE_PARAMS.AnaliseConfig
gerar_jogos_top10 = GERADOR_JOGOS.gerar_jogos_top10
exportar_jogos_txt = GERADOR_JOGOS.exportar_jogos_txt
validar_jogos_historico = VALIDADOR_1000_JOGOS.validar_jogos_historico
comparar_series = VALIDADOR_1000_JOGOS.comparar_series
gerar_relatorio_validacao = VALIDADOR_1000_JOGOS.gerar_relatorio_validacao
gerar_previsao_proximo_sorteio = PREVISAO_30N.gerar_previsao_proximo_sorteio
executar_ciclo_refinamento = CICLO_REFINAMENTO.executar_ciclo_refinamento
# Módulos adicionais
Analisador9N = ANALISADOR_9_NUMEROS
Indicador10N = INDICADOR_OTIMIZADO_10N
Analisador20N = ANALISADOR_UNIVERSO_REDUZIDO
SistemaExport = SISTEMA_EXPORTACAO

# Tentar importar colorama para cores
try:
    from colorama import init, Fore, Style
    init(autoreset=True)
    CORES_DISPONIVEIS = True
except ImportError:
    CORES_DISPONIVEIS = False
    # Fallback sem cores
    class Fore:
        GREEN = YELLOW = RED = CYAN = MAGENTA = BLUE = WHITE = ""
    class Style:
        BRIGHT = RESET_ALL = ""


def limpar_tela():
    """Limpa a tela do terminal."""
    os.system('cls' if os.name == 'nt' else 'clear')


def exibir_banner():
    """Exibe o banner do sistema."""
    banner = f"""
{Fore.YELLOW}
███╗   ███╗███████╗ ██████╗  █████╗  ██████╗██╗     ██╗
████╗ ████║██╔════╝██╔════╝ ██╔══██╗██╔════╝██║     ██║
██╔████╔██║█████╗  ██║  ███╗███████║██║     ██║     ██║
██║╚██╔╝██║██╔══╝  ██║   ██║██╔══██║██║     ██║     ██║
██║ ╚═╝ ██║███████╗╚██████╔╝██║  ██║╚██████╗███████╗██║
╚═╝     ╚═╝╚══════╝ ╚═════╝ ╚═╝  ╚═╝ ╚═════╝╚══════╝╚═╝
                                                 v6.2
{Style.RESET_ALL}
{Fore.CYAN}╔══════════════════════════════════════════════════════════════════╗
║                                                                  ║
║         {Fore.WHITE}Sistema Inteligente para Análise Mega-Sena{Fore.CYAN}           ║
║                                                                  ║
╚══════════════════════════════════════════════════════════════════╝{Style.RESET_ALL}
"""
    print(banner)


def exibir_menu_principal():
    """Exibe o menu principal."""
    menu = f"""
{Fore.GREEN}┌──────────────────────────────────────────────────────────────┐
│                     MENU PRINCIPAL                           │
└──────────────────────────────────────────────────────────────┘{Style.RESET_ALL}

{Fore.YELLOW}1.{Style.RESET_ALL} 🎯 Gerar 210 Jogos (Top 10 Indicadores)
{Fore.YELLOW}2.{Style.RESET_ALL} 📊 Validar Jogos com 1000 Sorteios Históricos
{Fore.YELLOW}3.{Style.RESET_ALL} 🚀 Análise Completa (Geração + Validação)
{Fore.YELLOW}4.{Style.RESET_ALL} 📈 Visualizar Estatísticas dos Indicadores
{Fore.YELLOW}5.{Style.RESET_ALL} 💾 Exportar Jogos para Arquivo
{Fore.YELLOW}6.{Style.RESET_ALL} ⚙️  Configurações do Sistema
{Fore.YELLOW}7.{Style.RESET_ALL} 🎲 Análise de Universo Reduzido (9N / 10N / 20N)
{Fore.YELLOW}8.{Style.RESET_ALL} 🚀 Análise v6 Completa (9N + 10N + 20N + Validação) ⭐
{Fore.YELLOW}9.{Style.RESET_ALL} 🔮 Previsão TOP_30N (Próximo Sorteio) 🆕
{Fore.YELLOW}10.{Style.RESET_ALL} 🤖 Ciclo de Refinamento IA (Novo)
{Fore.YELLOW}11.{Style.RESET_ALL} 🔄 Validação Retroativa e Auto-Aprendizado ⭐
{Fore.YELLOW}12.{Style.RESET_ALL} 🔒 Análise Conservadora (Anti-Overfitting) 🆕 v6.2
{Fore.RED}0.{Style.RESET_ALL} ❌ Sair

{Fore.CYAN}══════════════════════════════════════════════════════════════{Style.RESET_ALL}
"""
    print(menu)


def validar_entrada(prompt: str, opcoes_validas: list) -> int:
    """
    Valida entrada do usuário.
    
    Args:
        prompt: Mensagem a exibir
        opcoes_validas: Lista de opções válidas
        
    Returns:
        Opção escolhida ou None se muitas tentativas
    """
    tentativas = 0
    while True:
        if tentativas >= 3:
            print(f"{Fore.RED}❌ Muitas tentativas inválidas. Retornando ao menu.{Style.RESET_ALL}")
            return None
        
        try:
            entrada = input(prompt)
            opcao = int(entrada)
            
            if opcao in opcoes_validas:
                return opcao
            else:
                print(f"{Fore.RED}❌ Opção inválida! Escolha: {opcoes_validas}{Style.RESET_ALL}")
                tentativas += 1
        except ValueError:
            print(f"{Fore.RED}❌ Digite um número válido!{Style.RESET_ALL}")
            tentativas += 1
        except KeyboardInterrupt:
            print(f"\n{Fore.YELLOW}⚠️  Operação cancelada pelo usuário{Style.RESET_ALL}")
            return 0


def carregar_dados() -> Optional[pd.DataFrame]:
    """
    Carrega dados históricos.
    
    Returns:
        DataFrame com histórico ou None se erro
    """
    try:
        print(f"\n{Fore.CYAN}📂 Carregando dados históricos...{Style.RESET_ALL}")
        df = pd.read_excel(str(ARQUIVO_HISTORICO), sheet_name='MEGA SENA')
        print(f"{Fore.GREEN}✅ {len(df)} sorteios carregados com sucesso!{Style.RESET_ALL}")
        return df
    except Exception as e:
        print(f"{Fore.RED}❌ Erro ao carregar dados: {e}{Style.RESET_ALL}")
        return None


def opcao_1_gerar_jogos(df_historico: pd.DataFrame):
    """Opção 1: Gerar 210 jogos com top 10 indicadores."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"🎯 GERAÇÃO DE 210 JOGOS COM TOP 10 INDICADORES")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    # Carregar ranking existente ou criar um simples
    print(f"{Fore.CYAN}📊 Preparando ranking de indicadores...{Style.RESET_ALL}")
    
    # Importar módulos necessários
    try:
        # from validacao.ranking_indicadores import criar_ranking (Substituído por Fonte)
        # from validacao.analisador_historico import avaliar_serie_historica_completa
        # Usando Fontes já importadas no topo
        criar_ranking = RANKING.criar_ranking
        avaliar_serie_historica_completa = ANALISADOR_HISTORICO.avaliar_serie_historica_completa
        
        print(f"{Fore.CYAN}🔍 Executando análise histórica (isso pode levar alguns minutos)...{Style.RESET_ALL}")
        
        # Avaliar histórico
        estatisticas = avaliar_serie_historica_completa(
            df_historico,
            janela_inicial=len(df_historico) - AnaliseConfig.BATIMENTO_JANELA_OFFSET,
            passo=AnaliseConfig.BATIMENTO_PASSO,
            max_jogos=AnaliseConfig.BATIMENTO_MAX_JOGOS
        )
        
        # Criar ranking
        estats_dict = {nome: estat.to_dict() for nome, estat in estatisticas.items()}
        ranking = criar_ranking(estats_dict)
        
        print(f"{Fore.GREEN}✅ Ranking criado com {len(ranking)} indicadores{Style.RESET_ALL}")
        
        # --- ATUALIZAÇÃO DA PLANILHA (Solicitado pelo Usuário) ---
        print(f"{Fore.CYAN}💾 Salvando estatísticas atualizadas na planilha...{Style.RESET_ALL}")
        try:
            from validacao.ranking_indicadores import gerar_dataframe_ranking
            from openpyxl import load_workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # 1. Converter novo ranking para DF
            df_novo = gerar_dataframe_ranking(ranking)
            
            # 2. Carregar DF existente para preservar Peso_IA
            arquivo_excel = RESULTADO_DIR / 'ANALISE_HISTORICO_COMPLETO.xlsx'
            if os.path.exists(arquivo_excel):
                try:
                    df_antigo = pd.read_excel(arquivo_excel, sheet_name='RANKING INDICADORES')
                    
                    # Se tiver colunas essenciais, fazer merge
                    if 'Peso_Atual' in df_antigo.columns and 'indicador' in df_novo.columns:
                         # Renomear para facilitar merge
                         df_novo = df_novo.rename(columns={'indicador': 'Indicador'})
                         if 'Indicador' in df_antigo.columns:
                             # Merge mantendo Peso_Atual do antigo
                             df_merged = pd.merge(df_novo, df_antigo[['Indicador', 'Peso_Atual', 'Descrição', 'Categoria']], on='Indicador', how='left')
                             # Se Peso_Atual vier NaN (indicador novo), preencher
                             df_merged['Peso_Atual'] = df_merged['Peso_Atual'].fillna(50.0)
                             df_novo = df_merged
                except:
                    # Se der erro ao ler antigo, segue com o novo (renomeando)
                    df_novo = df_novo.rename(columns={'indicador': 'Indicador'})
            
            # 3. Salvar
            wb = load_workbook(arquivo_excel)
            if 'RANKING INDICADORES' in wb.sheetnames:
                ws = wb['RANKING INDICADORES']
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet('RANKING INDICADORES')
                
            for row in dataframe_to_rows(df_novo, index=False, header=True):
                ws.append(row)
                
            wb.save(arquivo_excel)
            print(f"{Fore.GREEN}   ✅ Aba RANKING INDICADORES atualizada!{Style.RESET_ALL}")
            
        except Exception as save_err:
             print(f"{Fore.YELLOW}⚠️  Aviso: Não foi possível salvar planilha de ranking: {save_err}{Style.RESET_ALL}")
        # ---------------------------------------------------------
        
    except Exception as e:
        print(f"{Fore.YELLOW}⚠️  Não foi possível carregar ranking completo: {e}{Style.RESET_ALL}")
        print(f"{Fore.CYAN}   Usando ranking simplificado...{Style.RESET_ALL}")
        
        # Ranking simplificado
        ranking = [
            {'indicador': f'Indicador{i}', 'relevancia': 100 - i*5, 'estrelas': '⭐'*3}
            for i in range(1, 43)
        ]
    
    # Gerar jogos
    jogos = gerar_jogos_top10(
        df_historico,
        ranking,
        n_jogos=AnaliseConfig.GERACAO_N_JOGOS,
        top_n=AnaliseConfig.GERACAO_TOP_INDICADORES,
        verbose=True
    )
    
    # Salvar jogos usando sistema centralizado
    # from src.utils.sistema_exportacao import SistemaExportacao, criar_config_jogos
    # Usando fonte
    SistemaExportacao = SistemaExport.SistemaExportacao
    criar_config_jogos = SistemaExport.criar_config_jogos
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sistema = SistemaExportacao()
    
    # Exportar para TXT
    config_txt = {
        "tipo": "txt",
        "arquivo": str(RESULTADO_DIR / f"jogos_gerados_{timestamp}.txt"),
        "dados": jogos,
        "formatacao": {
            "titulo": "MEGACLI v6.0 - JOGOS GERADOS COM TOP 10 INDICADORES"
        },
        "formato_linha": "#{rank:03d}: {numeros_str} | Score: {score:.2f} | Prob: {probabilidade:.1f}% | {confianca}"
    }
    
    # Preparar dados para TXT
    dados_txt = []
    for jogo in jogos:
        nums_str = '-'.join(f"{n:02d}" for n in jogo['numeros'])
        dados_txt.append({
            'rank': jogo['rank'],
            'numeros_str': nums_str,
            'score': jogo['score'],
            'probabilidade': jogo['probabilidade'],
            'confianca': jogo['confianca']
        })
    config_txt['dados'] = dados_txt
    
    # Exportar TXT
    sistema.exportar(config_txt)
    
    # Exportar para Excel
    config_excel = criar_config_jogos(
        arquivo=str(RESULTADO_DIR / 'ANALISE_HISTORICO_COMPLETO.xlsx'),
        jogos=jogos,
        aba=f"JOGOS_{timestamp[:8]}"
    )
    sistema.exportar(config_excel)
    
    print(f"\n{Fore.GREEN}✅ Jogos salvos em TXT e Excel{Style.RESET_ALL}")
    
    # Armazenar jogos globalmente para outras opções
    global JOGOS_GERADOS
    JOGOS_GERADOS = jogos
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_2_validar_jogos(df_historico: pd.DataFrame):
    """Opção 2: Validar jogos com 1000 sorteios."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"📊 VALIDAÇÃO COM 1000 SORTEIOS HISTÓRICOS")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    # Verificar se há jogos gerados
    global JOGOS_GERADOS
    if JOGOS_GERADOS is None:
        print(f"{Fore.YELLOW}⚠️  Nenhum jogo gerado ainda!{Style.RESET_ALL}")
        print(f"{Fore.CYAN}   Por favor, execute a Opção 1 primeiro.{Style.RESET_ALL}")
        input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        return
    
    # Validar jogos
    validacao = validar_jogos_historico(
        JOGOS_GERADOS,
        df_historico,
        n_sorteios=AnaliseConfig.VALIDACAO_N_SORTEIOS,
        splits=AnaliseConfig.VALIDACAO_SPLIT_SERIES,
        verbose=True
    )
    
    # Comparar séries
    print(f"\n{Fore.CYAN}📊 Comparando performance entre séries...{Style.RESET_ALL}")
    df_comparacao = comparar_series(validacao['resultados'])
    
    print(f"\n{Fore.GREEN}{'='*70}")
    print(f"TOP 10 JOGOS - COMPARAÇÃO ENTRE SÉRIES")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    print(df_comparacao.head(10).to_string(index=False))
    
    # Salvar relatório
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    arquivo_json = RESULTADO_DIR / f"validacao_{timestamp}.json"
    gerar_relatorio_validacao(validacao, str(arquivo_json))
    
    print(f"\n{Fore.GREEN}✅ Relatório salvo em: {arquivo_json}{Style.RESET_ALL}")
    
    # Armazenar validação globalmente
    global VALIDACAO_ATUAL
    VALIDACAO_ATUAL = validacao
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_3_analise_completa(df_historico: pd.DataFrame):
    """Opção 3: Análise completa (geração + validação)."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"🚀 ANÁLISE COMPLETA (GERAÇÃO + VALIDAÇÃO)")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    print(f"{Fore.CYAN}Esta operação executará:{Style.RESET_ALL}")
    print(f"  1. Geração de 210 jogos com top 10 indicadores")
    print(f"  2. Validação com 1000 sorteios históricos")
    print(f"\n{Fore.YELLOW}⏱️  Tempo estimado: 15-20 minutos{Style.RESET_ALL}\n")
    
    confirmar = input(f"{Fore.CYAN}Deseja continuar? (s/n): {Style.RESET_ALL}").lower()
    
    if confirmar != 's':
        print(f"{Fore.YELLOW}⚠️  Operação cancelada{Style.RESET_ALL}")
        input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        return
    
    # Executar geração
    print(f"\n{Fore.GREEN}{'='*70}")
    print(f"ETAPA 1/2: GERAÇÃO DE JOGOS")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    opcao_1_gerar_jogos(df_historico)
    
    # Executar validação
    print(f"\n{Fore.GREEN}{'='*70}")
    print(f"ETAPA 2/2: VALIDAÇÃO HISTÓRICA")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    opcao_2_validar_jogos(df_historico)
    
    print(f"\n{Fore.GREEN}✅ ANÁLISE COMPLETA FINALIZADA!{Style.RESET_ALL}")
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_4_visualizar_estatisticas():
    """Opção 4: Visualizar estatísticas."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"📈 ESTATÍSTICAS DOS INDICADORES")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    # Exibir configurações atuais
    AnaliseConfig.exibir_configuracao()
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_5_exportar_jogos():
    """Opção 5: Exportar jogos."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"💾 EXPORTAR JOGOS PARA ARQUIVO")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    global JOGOS_GERADOS
    if JOGOS_GERADOS is None:
        print(f"{Fore.YELLOW}⚠️  Nenhum jogo gerado ainda!{Style.RESET_ALL}")
        print(f"{Fore.CYAN}   Por favor, execute a Opção 1 primeiro.{Style.RESET_ALL}")
        input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        return
    
    print(f"{Fore.CYAN}Total de jogos disponíveis: {len(JOGOS_GERADOS)}{Style.RESET_ALL}\n")
    
    # Escolher formato
    print(f"{Fore.GREEN}Formatos disponíveis:{Style.RESET_ALL}")
    print(f"  1. TXT (texto formatado)")
    print(f"  2. CSV (planilha)")
    print(f"  0. Cancelar\n")
    
    formato = validar_entrada("Escolha o formato: ", [0, 1, 2])
    
    if formato == 0:
        return
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    if formato == 1:
        arquivo = RESULTADO_DIR / f"jogos_export_{timestamp}.txt"
        exportar_jogos_txt(JOGOS_GERADOS, str(arquivo))
        print(f"\n{Fore.GREEN}✅ Jogos exportados para: {arquivo}{Style.RESET_ALL}")
    
    elif formato == 2:
        arquivo = RESULTADO_DIR / f"jogos_export_{timestamp}.csv"
        df = pd.DataFrame([
            {
                'Rank': j['rank'],
                'Numeros': '-'.join(f"{n:02d}" for n in j['numeros']),
                'Score': j['score'],
                'Probabilidade': j['probabilidade'],
                'Confianca': j['confianca']
            }
            for j in JOGOS_GERADOS
        ])
        df.to_csv(arquivo, index=False, encoding='utf-8-sig')
        print(f"\n{Fore.GREEN}✅ Jogos exportados para: {arquivo}{Style.RESET_ALL}")
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_6_configuracoes():
    """Opção 6: Configurações."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"⚙️  CONFIGURAÇÕES DO SISTEMA")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    # Exibir configurações
    AnaliseConfig.exibir_configuracao()
    
    # Validar parâmetros
    print(f"\n{Fore.CYAN}🔍 Validando parâmetros...{Style.RESET_ALL}")
    AnaliseConfig.validar_parametros()
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_7_universo_reduzido(df_historico: pd.DataFrame):
    """Opção 7: Análise de Universo Reduzido (9N, 10N ou 20N)."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"🎲 ANÁLISE DE UNIVERSO REDUZIDO")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    # Submenu de seleção de universo
    print(f"{Fore.GREEN}┌──────────────────────────────────────────────────────────────┐")
    print(f"│         ESCOLHA O TAMANHO DO UNIVERSO REDUZIDO               │")
    print(f"└──────────────────────────────────────────────────────────────┘{Style.RESET_ALL}\n")
    
    print(f"{Fore.YELLOW}1.{Style.RESET_ALL} 🎯 TOP 9N  - 9 números  (84 jogos,   R$ 420,00)")
    print(f"{Fore.YELLOW}2.{Style.RESET_ALL} 🎲 TOP 10N - 10 números (210 jogos,  R$ 1.050,00)")
    print(f"{Fore.YELLOW}3.{Style.RESET_ALL} 📊 TOP 20N - 20 números (38.760 jogos, R$ 193.800,00)")
    print(f"{Fore.RED}0.{Style.RESET_ALL} ⬅️  Voltar ao menu principal\n")
    
    print(f"{Fore.CYAN}═══════════════════════════════════════════════════════════════{Style.RESET_ALL}")
    
    escolha = validar_entrada("Escolha o universo: ", [0, 1, 2, 3])
    
    if escolha == 0:
        return
    
    # Executar análise baseada na escolha
    if escolha == 1:
        opcao_7_1_top_9n(df_historico)
    elif escolha == 2:
        opcao_7_2_top_10n(df_historico)
    elif escolha == 3:
        opcao_7_3_top_20n(df_historico)


def opcao_7_1_top_9n(df_historico: pd.DataFrame):
    """Análise TOP 9N."""
    # Imports via Fonte
    selecionar_top_9_numeros = Analisador9N.selecionar_top_9_numeros
    analisar_combinacoes_9 = Analisador9N.analisar_combinacoes_9
    validar_cobertura_9 = Analisador9N.validar_cobertura_9
    gerar_todos_jogos_9 = Analisador9N.gerar_todos_jogos_9
    
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"🎯 ANÁLISE TOP 9N (9 NÚMEROS)")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    # Ranking de teste
    ranking_teste = [{'indicador': f'Ind{i}', 'relevancia': 100-i*5} for i in range(1, 11)]
    
    # Selecionar top 9
    numeros_9, scores_9 = selecionar_top_9_numeros(df_historico, ranking_teste, verbose=True)
    
    # Analisar combinações
    analise_9 = analisar_combinacoes_9(numeros_9, verbose=True)
    
    # Validar cobertura
    validacao_9 = validar_cobertura_9(numeros_9, df_historico, verbose=True)
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_7_2_top_10n(df_historico: pd.DataFrame):
    """Análise TOP 10N."""
    # Imports via Fonte
    IndicadorOtimizado10N = Indicador10N.IndicadorOtimizado10N
    import math
    
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"🎲 ANÁLISE TOP 10N (10 NÚMEROS)")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    # Criar indicador otimizado
    indicador = IndicadorOtimizado10N()
    
    # Selecionar top 10
    numeros_10, scores_10 = indicador.selecionar_top_10(df_historico, verbose=True)
    
    # Analisar combinações
    print(f"\n📊 Análise de Combinações")
    print("="*70)
    comb_10 = math.comb(10, 6)
    print(f"   • Total de combinações: {comb_10:,}")
    print(f"   • Custo (R$ 5,00/jogo): R$ {comb_10 * 5.0:,.2f}")
    
    # Validar cobertura
    validacao = indicador.validar_cobertura(numeros_10, df_historico, verbose=True)
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_7_3_top_20n(df_historico: pd.DataFrame):
    """Análise TOP 20N."""
    # Imports via Fonte
    selecionar_top_20_numeros = Analisador20N.selecionar_top_20_numeros
    analisar_combinacoes = Analisador20N.analisar_combinacoes
    gerar_estrategias_cobertura = Analisador20N.gerar_estrategias_cobertura
    gerar_jogos_universo_reduzido = Analisador20N.gerar_jogos_universo_reduzido
    analisar_probabilidades = Analisador20N.analisar_probabilidades
    
    # Variáveis locais para armazenar dados
    numeros_20 = None
    pesos_20 = None
    
    while True:
        limpar_tela()
        exibir_banner()
        
        print(f"\n{Fore.YELLOW}{'='*70}")
        print(f"📊 ANÁLISE TOP 20N (20 NÚMEROS)")
        print(f"{'='*70}{Style.RESET_ALL}\n")
        
        # Submenu
        print(f"{Fore.GREEN}┌──────────────────────────────────────────────────────────────┐")
        print(f"│              ANÁLISE DE UNIVERSO REDUZIDO                    │")
        print(f"└──────────────────────────────────────────────────────────────┘{Style.RESET_ALL}\n")
        
        print(f"{Fore.YELLOW}1.{Style.RESET_ALL} 📊 Selecionar os 20 números mais prováveis")
        print(f"{Fore.YELLOW}2.{Style.RESET_ALL} 🔢 Analisar combinações possíveis")
        print(f"{Fore.YELLOW}3.{Style.RESET_ALL} 🎯 Ver estratégias de cobertura")
        print(f"{Fore.YELLOW}4.{Style.RESET_ALL} 🎲 Gerar jogos do universo reduzido")
        print(f"{Fore.YELLOW}5.{Style.RESET_ALL} 📈 Análise de probabilidades")
        print(f"{Fore.RED}0.{Style.RESET_ALL} ⬅️  Voltar ao menu anterior\n")
        
        print(f"{Fore.CYAN}═══════════════════════════════════════════════════════════════{Style.RESET_ALL}")
        
        opcao = validar_entrada("Escolha uma opção: ", [0, 1, 2, 3, 4, 5])
        
        if opcao == 0:
            break
        
        elif opcao == 1:
            # Selecionar top 20 números
            print(f"\n{Fore.CYAN}📊 Preparando ranking de indicadores...{Style.RESET_ALL}")
            
            try:
                from validacao.ranking_indicadores import criar_ranking
                from validacao.analisador_historico import avaliar_serie_historica_completa
                
                # Avaliar histórico
                estatisticas = avaliar_serie_historica_completa(
                    df_historico,
                    janela_inicial=len(df_historico) - AnaliseConfig.BATIMENTO_JANELA_OFFSET,
                    passo=AnaliseConfig.BATIMENTO_PASSO,
                    max_jogos=AnaliseConfig.BATIMENTO_MAX_JOGOS
                )
                
                estats_dict = {nome: estat.to_dict() for nome, estat in estatisticas.items()}
                ranking = criar_ranking(estats_dict)
                
            except Exception as e:
                print(f"{Fore.YELLOW}⚠️  Usando ranking simplificado{Style.RESET_ALL}")
                ranking = [
                    {'indicador': f'Ind{i}', 'relevancia': 100-i*5}
                    for i in range(1, 11)
                ]
            
            # Selecionar top 20
            numeros_20, pesos_20 = selecionar_top_20_numeros(
                df_historico,
                ranking,
                top_indicadores=10,
                janela=100,
                verbose=True
            )
            
            input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        
        elif opcao == 2:
            # Analisar combinações
            if numeros_20 is None:
                print(f"\n{Fore.YELLOW}⚠️  Execute a Opção 1 primeiro!{Style.RESET_ALL}")
                input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
                continue
            
            analisar_combinacoes(numeros_20, verbose=True)
            input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        
        elif opcao == 3:
            # Estratégias de cobertura
            if numeros_20 is None:
                print(f"\n{Fore.YELLOW}⚠️  Execute a Opção 1 primeiro!{Style.RESET_ALL}")
                input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
                continue
            
            gerar_estrategias_cobertura(numeros_20, verbose=True)
            input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        
        elif opcao == 4:
            # Gerar jogos
            if numeros_20 is None or pesos_20 is None:
                print(f"\n{Fore.YELLOW}⚠️  Execute a Opção 1 primeiro!{Style.RESET_ALL}")
                input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
                continue
            
            print(f"\n{Fore.CYAN}Quantos jogos deseja gerar?{Style.RESET_ALL}")
            print(f"  Sugestões:")
            print(f"  • 388 jogos (Top 1% - R$ 1.940,00)")
            print(f"  • 1.938 jogos (Top 5% - R$ 9.690,00)")
            print(f"  • 38.760 jogos (Cobertura total - R$ 193.800,00)\n")
            
            try:
                n_jogos = int(input(f"{Fore.CYAN}Número de jogos: {Style.RESET_ALL}"))
                
                jogos = gerar_jogos_universo_reduzido(
                    numeros_20,
                    pesos_20,
                    n_jogos=n_jogos,
                    estrategia='otimizada',
                    verbose=True
                )
                
                # Salvar
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                arquivo = RESULTADO_DIR / f"universo_reduzido_{timestamp}.txt"
                
                with open(arquivo, 'w', encoding='utf-8') as f:
                    f.write(f"UNIVERSO REDUZIDO - {n_jogos} JOGOS\n")
                    f.write("="*70 + "\n\n")
                    f.write(f"20 Números: {'-'.join(f'{n:02d}' for n in numeros_20)}\n\n")
                    f.write("JOGOS GERADOS:\n")
                    f.write("="*70 + "\n")
                    for i, jogo in enumerate(jogos, 1):
                        f.write(f"#{i:04d}: {'-'.join(f'{n:02d}' for n in jogo)}\n")
                
                print(f"\n{Fore.GREEN}✅ Jogos salvos em: {arquivo}{Style.RESET_ALL}")
                
            except ValueError:
                print(f"\n{Fore.RED}❌ Número inválido!{Style.RESET_ALL}")
            
            input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        
        elif opcao == 5:
            # Análise de probabilidades
            if numeros_20 is None:
                print(f"\n{Fore.YELLOW}⚠️  Execute a Opção 1 primeiro!{Style.RESET_ALL}")
                input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
                continue
            
            analisar_probabilidades(numeros_20, df_historico, janela=100, verbose=True)
            input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_8_analise_v6(df_historico: pd.DataFrame):
    """Opção 8: Análise v6 Completa (9N + 10N + 20N + Validação)."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"🚀 ANÁLISE V6 COMPLETA (UNIVERSOS REDUZIDOS + VALIDAÇÃO)")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    print(f"{Fore.CYAN}Esta operação executará:{Style.RESET_ALL}")
    print(f"  1. Análise histórica completa (BATIMENTO + Eficácias)")
    print(f"  2. Ranking de indicadores")
    print(f"  3. Validação IA")
    print(f"  4. Seleção de universo reduzido (20 números)")
    print(f"  5. Análise adicional (10 números) ⭐")
    print(f"  6. Análise adicional (9 números)")
    print(f"  7. Geração de 210 jogos")
    print(f"  8. Validação histórica (100 jogos)")
    print(f"  9. Criação de abas TOP_9N, TOP_10N e TOP_20N no Excel")
    print(f"\n{Fore.YELLOW}⏱️  Tempo estimado: 25-30 minutos{Style.RESET_ALL}\n")
    
    confirmar = input(f"{Fore.CYAN}Deseja continuar? (s/n): {Style.RESET_ALL}").lower()
    
    if confirmar != 's':
        print(f"{Fore.YELLOW}⚠️  Operação cancelada{Style.RESET_ALL}")
        input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        return
    
    # Executar gerar_analise_v6.py
    print(f"\n{Fore.GREEN}{'='*70}")
    print(f"EXECUTANDO ANÁLISE V6 COMPLETA")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    try:
        # Executar script v6 via Fonte
        ANALISE_V6.main()
    
    except Exception as e:
        print(f"\n{Fore.RED}❌ Erro ao executar análise v6: {e}{Style.RESET_ALL}")
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_9_previsao_30n(df_historico: pd.DataFrame):
    """Opção 9: Previsão TOP_30N para próximo sorteio."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.YELLOW}{'='*70}")
    print(f"🔮 PREVISÃO TOP_30N - PRÓXIMO SORTEIO")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    print(f"{Fore.CYAN}Esta operação executará:{Style.RESET_ALL}")
    print(f"  1. Análise dos últimos 500 jogos históricos")
    print(f"  2. Seleção dos 30 números mais prováveis")
    print(f"  3. Refinamento para os 10 melhores números")
    print(f"  4. Geração de previsão para o próximo sorteio")
    print(f"\n{Fore.YELLOW}⏱️  Tempo estimado: 2-3 minutos{Style.RESET_ALL}\n")
    
    confirmar = input(f"{Fore.CYAN}Deseja continuar? (s/n): {Style.RESET_ALL}").lower()
    
    if confirmar != 's':
        print(f"\n{Fore.YELLOW}⚠️  Operação cancelada{Style.RESET_ALL}")
        input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        return
    
    try:
        # Executar previsão
        print(f"\n{Fore.CYAN}🔄 Iniciando análise preditiva...{Style.RESET_ALL}\n")
        resultado = gerar_previsao_proximo_sorteio(df_historico)
        
        if resultado:
            print(f"\n{Fore.GREEN}{'='*70}")
            print(f"✅ PREVISÃO GERADA COM SUCESSO!")
            print(f"{'='*70}{Style.RESET_ALL}\n")
            
            print(f"{Fore.CYAN}📊 Resumo da Análise:{Style.RESET_ALL}")
            print(f"  • Jogos analisados: {resultado.get('jogos_analisados', 500)}")
            print(f"  • Números selecionados (TOP_30): {len(resultado.get('top_30', []))}")
            print(f"  • Números refinados (TOP_10): {len(resultado.get('top_10', []))}")
            print(f"  • Confiança: {resultado.get('confianca', 'N/A')}")
            
            print(f"\n{Fore.YELLOW}🎯 APOSTA RECOMENDADA:{Style.RESET_ALL}")
            aposta = resultado.get('aposta_final', [])
            if aposta:
                aposta_str = '-'.join(f"{n:02d}" for n in sorted(aposta))
                print(f"  {Fore.GREEN}{aposta_str}{Style.RESET_ALL}")
            
            print(f"\n{Fore.CYAN}💾 Resultados salvos em:{Style.RESET_ALL}")
            print(f"  • {RESULTADO_DIR / 'ANALISE_HISTORICO_COMPLETO.xlsx'}")
            print(f"    Aba: {Fore.YELLOW}PREVISÕES{Style.RESET_ALL}")
            
        else:
            print(f"\n{Fore.RED}❌ Erro ao gerar previsão{Style.RESET_ALL}")
            
    except Exception as e:
        print(f"\n{Fore.RED}❌ Erro durante previsão: {e}{Style.RESET_ALL}")
        import traceback
        traceback.print_exc()
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_10_refinamento_ia(df_historico: pd.DataFrame):
    """Executa o ciclo de refinamento iterativo com IA."""
    print(f"\n{Fore.CYAN}🤖 Iniciando Ciclo de Refinamento IA...{Style.RESET_ALL}")
    
    from src.core.ciclo_refinamento_ia import executar_ciclo_refinamento
    import os
    
    # Verificar API Key
    if not os.environ.get("GOOGLE_API_KEY"):
        print(f"\n{Fore.RED}❌ Erro: GOOGLE_API_KEY não configurada no arquivo .env{Style.RESET_ALL}")
        return

    try:
        # Reload para garantir dados frescos (Hot Reload)
        print(f"\n{Fore.CYAN}🔄 Recarregando dados históricos para garantir atualização...{Style.RESET_ALL}")
        df_fresh = carregar_dados()
        if df_fresh is not None:
             executar_ciclo_refinamento(df_fresh)
        else:
             print(f"{Fore.YELLOW}⚠️  Falha ao recarregar. Usando dados em memória.{Style.RESET_ALL}")
             executar_ciclo_refinamento(df_historico)
        input(f"\n{Fore.GREEN}Pressione ENTER para voltar ao menu...{Style.RESET_ALL}")
    except Exception as e:
        print(f"\n{Fore.RED}❌ Erro durante o ciclo: {e}{Style.RESET_ALL}")
        import traceback
        traceback.print_exc()
        input("\nPressione ENTER para continuar...")


def opcao_11_validacao_retroativa(df_historico: pd.DataFrame):
    """Opção 11: Validação Retroativa e Auto-Aprendizado."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.CYAN}{'='*70}")
    print(f"OPÇÃO 11: VALIDAÇÃO RETROATIVA E AUTO-APRENDIZADO")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    print(f"{Fore.YELLOW}🔍 Esta função irá:{Style.RESET_ALL}\n")
    print(f"   1. Ler últimos sorteios da Mega-Sena (Dados/)")
    print(f"   2. Comparar com previsões geradas (Resultado/ANALISE_HISTORICO_COMPLETO.xlsx)")
    print(f"   3. Calcular taxa de acerto (6, 5, 4, 3 números)")
    print(f"   4. Analisar contribuição de cada indicador")
    print(f"   5. Ajustar pesos dos indicadores automaticamente")
    print(f"   6. Gerar relatório detalhado de performance\n")
    
    print(f"{Fore.CYAN}📊 Configuração:{Style.RESET_ALL}")
    print(f"   • Taxa de aprendizado: 10% (moderada)")
    print(f"   • Sorteios analisados: 5 últimos (configurável)")
    print(f"   • Critério de sucesso: 4+ acertos\n")
    
    confirmar = input(f"{Fore.CYAN}Deseja continuar? (s/n): {Style.RESET_ALL}").strip().lower()
    
    if confirmar != 's':
        print(f"\n{Fore.YELLOW}⚠️  Operação cancelada{Style.RESET_ALL}")
        input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        return
    
    # Perguntar quantos sorteios analisar
    print(f"\n{Fore.CYAN}Quantos sorteios deseja analisar?{Style.RESET_ALL}")
    print(f"   Recomendação: 5 para análise rápida, 10-20 para análise completa")
    
    try:
        n_sorteios = int(input(f"{Fore.CYAN}Número de sorteios (padrão: 5): {Style.RESET_ALL}") or "5")
        n_sorteios = max(1, min(50, n_sorteios))  # Limitar entre 1 e 50
    except ValueError:
        n_sorteios = 5
        print(f"{Fore.YELLOW}⚠️  Valor inválido, usando padrão: 5{Style.RESET_ALL}")
    
    # Executar validação
    try:
        print(f"\n{Fore.CYAN}⚙️  Executando validação retroativa...{Style.RESET_ALL}\n")
        
        from src.validacao.validador_retroativo import executar_validacao_completa
        
        resultado = executar_validacao_completa(n_ultimos_sorteios=n_sorteios)
        
        if resultado:
            print(f"\n{Fore.GREEN}{'='*70}")
            print(f"✅ VALIDAÇÃO CONCLUÍDA COM SUCESSO!")
            print(f"{'='*70}{Style.RESET_ALL}")
            
            # Mostrar opções adicionais
            print(f"\n{Fore.CYAN}📁 Arquivos gerados:{Style.RESET_ALL}")
            print(f"   • Análise completa: Resultado/validacao_retroativa/")
            print(f"   • Relatório TXT com detalhes")
            print(f"   • Análise JSON com dados brutos\n")
            
            print(f"{Fore.YELLOW}💡 Dica:{Style.RESET_ALL}")
            print(f"   Os pesos ajustados serão aplicados automaticamente")
            print(f"   na próxima execução do sistema.\n")
        else:
            print(f"\n{Fore.RED}{'='*70}")
            print(f"❌ ERRO NA VALIDAÇÃO")
            print(f"{'='*70}{Style.RESET_ALL}")
            print(f"\n{Fore.YELLOW}Verifique:{Style.RESET_ALL}")
            print(f"   • Se há arquivos em Dados/")
            print(f"   • Se ANALISE_HISTORICO_COMPLETO.xlsx existe em Resultado/")
            print(f"   • Se as abas PREVISÕES e RANKING INDICADORES existem")
    
    except Exception as e:
        print(f"\n{Fore.RED}❌ Erro durante validação retroativa: {e}{Style.RESET_ALL}\n")
        import traceback
        traceback.print_exc()
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def opcao_12_analise_conservadora(df_historico: pd.DataFrame):
    """Opção 12: Análise Conservadora com Anti-Overfitting."""
    limpar_tela()
    exibir_banner()
    
    print(f"\n{Fore.CYAN}{'='*70}")
    print(f"OPÇÃO 12: ANÁLISE CONSERVADORA (ANTI-OVERFITTING) v6.2")
    print(f"{'='*70}{Style.RESET_ALL}\n")
    
    print(f"{Fore.GREEN}🔒 Modo Conservador Estatisticamente Robusto:{Style.RESET_ALL}\n")
    print(f"   ✅ Usa apenas 5-7 indicadores mais robustos")
    print(f"   ✅ Universo mínimo de 25 números (reduz risco)")
    print(f"   ✅ Validação rigorosa Train/Test (80/20)")
    print(f"   ✅ Detecção automática de overfitting")
    print(f"   ✅ Intervalos de confiança em todas as métricas")
    print(f"   ✅ Gera 100 jogos (menor custo)")
    
    print(f"\n{Fore.YELLOW}📊 O que será executado:{Style.RESET_ALL}\n")
    print(f"   1. Filtrar indicadores mais robustos")
    print(f"   2. Validar com Split Train/Test (80% treino, 20% teste)")
    print(f"   3. Analisar overfitting (múltiplos critérios)")
    print(f"   4. Selecionar universo conservador (25 números)")
    print(f"   5. Gerar 100 jogos conservadores")
    print(f"   6. Relatório com intervalos de confiança\n")
    
    confirmar = input(f"{Fore.CYAN}Executar análise conservadora? (s/n): {Style.RESET_ALL}").strip().lower()
    
    if confirmar != 's':
        print(f"\n{Fore.YELLOW}⚠️  Operação cancelada{Style.RESET_ALL}")
        input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        return
    
    try:
        # 1. Criar ranking primeiro
        print(f"\n{Fore.CYAN}📊 Criando ranking de indicadores...{Style.RESET_ALL}\n")
        
        from validacao.analisador_historico import avaliar_serie_historica_completa
        from validacao.ranking_indicadores import criar_ranking
        
        estatisticas = avaliar_serie_historica_completa(
            df_historico,
            janela_inicial=len(df_historico) - AnaliseConfig.BATIMENTO_JANELA_OFFSET,
            passo=AnaliseConfig.BATIMENTO_PASSO,
            max_jogos=AnaliseConfig.BATIMENTO_MAX_JOGOS
        )
        estats_dict = {nome: estat.to_dict() for nome, estat in estatisticas.items()}
        ranking = criar_ranking(estats_dict)
        
        print(f"   ✅ Ranking criado com {len(ranking)} indicadores\n")
        
        # 2. Executar modo conservador
        from src.core.modo_conservador import ModoConservador
        
        modo_conserv = ModoConservador()
        resultado = modo_conserv.executar_analise_conservadora(df_historico, ranking)
        
        # 3. Salvar resultados
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_json = RESULTADO_DIR / f"analise_conservadora_{timestamp}.json"
        
        import json
        with open(arquivo_json, 'w', encoding='utf-8') as f:
            json.dump(resultado, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n💾 Análise salva em: {arquivo_json.name}")
        
        # 4. Salvar jogos
        arquivo_jogos = RESULTADO_DIR / f"jogos_conservadores_{timestamp}.txt"
        exportar_jogos_txt(resultado['jogos'], str(arquivo_jogos))
        
        print(f"💾 Jogos salvos em: {arquivo_jogos.name}")
        
        print(f"\n{Fore.GREEN}{'='*70}")
        print(f"✅ ANÁLISE CONSERVADORA CONCLUÍDA!")
        print(f"{'='*70}{Style.RESET_ALL}\n")
        
    except Exception as e:
        print(f"\n{Fore.RED}❌ Erro: {e}{Style.RESET_ALL}\n")
        import traceback
        traceback.print_exc()
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
    
    print(f"{Fore.GREEN}🔒 Modo Conservador Estatisticamente Robusto:{Style.RESET_ALL}\n")
    print(f"   ✅ Usa apenas 5-7 indicadores mais robustos")
    print(f"   ✅ Universo mínimo de 25 números (reduz risco)")
    print(f"   ✅ Validação rigorosa Train/Test (80/20)")
    print(f"   ✅ Detecção automática de overfitting")
    print(f"   ✅ Intervalos de confiança em todas as métricas")
    print(f"   ✅ Gera 100 jogos (menor custo)")
    
    print(f"\n{Fore.YELLOW}📊 O que será executado:{Style.RESET_ALL}\n")
    print(f"   1. Filtrar indicadores mais robustos")
    print(f"   2. Validar com Split Train/Test (80% treino, 20% teste)")
    print(f"   3. Analisar overfitting (múltiplos critérios)")
    print(f"   4. Selecionar universo conservador (25 números)")
    print(f"   5. Gerar 100 jogos conservadores")
    print(f"   6. Relatório com intervalos de confiança\n")
    
    confirmar = input(f"{Fore.CYAN}Executar análise conservadora? (s/n): {Style.RESET_ALL}").strip().lower()
    
    if confirmar != 's':
        print(f"\n{Fore.YELLOW}⚠️  Operação cancelada{Style.RESET_ALL}")
        input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")
        return
    
    try:
        # 1. Criar ranking primeiro
        print(f"\n{Fore.CYAN}📊 Criando ranking de indicadores...{Style.RESET_ALL}\n")
        
        from validacao.analisador_historico import avaliar_serie_historica_completa
        from validacao.ranking_indicadores import criar_ranking
        
        estatisticas = avaliar_serie_historica_completa(
            df_historico,
            janela_inicial=len(df_historico) - AnaliseConfig.BATIMENTO_JANELA_OFFSET,
            passo=AnaliseConfig.BATIMENTO_PASSO,
            max_jogos=AnaliseConfig.BATIMENTO_MAX_JOGOS
        )
        estats_dict = {nome: estat.to_dict() for nome, estat in estatisticas.items()}
        ranking = criar_ranking(estats_dict)
        
        print(f"   ✅ Ranking criado com {len(ranking)} indicadores\n")
        
        # 2. Executar modo conservador
        from src.core.modo_conservador import ModoConservador
        
        modo_conserv = ModoConservador()
        resultado = modo_conserv.executar_analise_conservadora(df_historico, ranking)
        
        # 3. Salvar resultados
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        arquivo_json = RESULTADO_DIR / f"analise_conservadora_{timestamp}.json"
        
        import json
        with open(arquivo_json, 'w', encoding='utf-8') as f:
            json.dump(resultado, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n💾 Análise salva em: {arquivo_json.name}")
        
        # 4. Salvar jogos
        arquivo_jogos = RESULTADO_DIR / f"jogos_conservadores_{timestamp}.txt"
        exportar_jogos_txt(resultado['jogos'], str(arquivo_jogos))
        
        print(f"💾 Jogos salvos em: {arquivo_jogos.name}")
        
        print(f"\n{Fore.GREEN}{'='*70}")
        print(f"✅ ANÁLISE CONSERVADORA CONCLUÍDA!")
        print(f"{'='*70}{Style.RESET_ALL}\n")
        
    except Exception as e:
        print(f"\n{Fore.RED}❌ Erro: {e}{Style.RESET_ALL}\n")
        import traceback
        traceback.print_exc()
    
    input(f"\n{Fore.CYAN}Pressione ENTER para continuar...{Style.RESET_ALL}")


def executar_menu():
    """Loop principal do menu."""
    # Variáveis globais para armazenar dados
    global JOGOS_GERADOS, VALIDACAO_ATUAL
    JOGOS_GERADOS = None
    VALIDACAO_ATUAL = None
    
    # Carregar dados históricos uma vez
    df_historico = carregar_dados()
    
    if df_historico is None:
        print(f"{Fore.RED}❌ Não foi possível carregar os dados. Encerrando...{Style.RESET_ALL}")
        return
    
    while True:
        limpar_tela()
        exibir_banner()
        exibir_menu_principal()
        
        opcao = validar_entrada("Escolha uma opção: ", [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12])
        
        if opcao == 0:
            print(f"\n{Fore.YELLOW}👋 Encerrando MegaCLI... Até logo!{Style.RESET_ALL}\n")
            break
        
        elif opcao == 1:
            opcao_1_gerar_jogos(df_historico)
        
        elif opcao == 2:
            opcao_2_validar_jogos(df_historico)
        
        elif opcao == 3:
            opcao_3_analise_completa(df_historico)
        
        elif opcao == 4:
            opcao_4_visualizar_estatisticas()
        
        elif opcao == 5:
            opcao_5_exportar_jogos()
        
        elif opcao == 6:
            opcao_6_configuracoes()
        
        elif opcao == 7:
            opcao_7_universo_reduzido(df_historico)
        
        elif opcao == 8:
            opcao_8_analise_v6(df_historico)
        
        elif opcao == 9:
            opcao_9_previsao_30n(df_historico)
        
        elif opcao == 10:
            opcao_10_refinamento_ia(df_historico)
        
        elif opcao == 11:
            opcao_11_validacao_retroativa(df_historico)
        
        elif opcao == 12:
            opcao_12_analise_conservadora(df_historico)


        
        elif opcao == 12:
            opcao_12_analise_conservadora(df_historico)
# Variáveis globais
JOGOS_GERADOS = None
VALIDACAO_ATUAL = None


if __name__ == "__main__":
    try:
        executar_menu()
    except KeyboardInterrupt:
        print(f"\n\n{Fore.YELLOW}⚠️  Programa interrompido pelo usuário{Style.RESET_ALL}\n")
    except Exception as e:
        print(f"\n{Fore.RED}❌ Erro inesperado: {e}{Style.RESET_ALL}\n")
        import traceback
        traceback.print_exc()
