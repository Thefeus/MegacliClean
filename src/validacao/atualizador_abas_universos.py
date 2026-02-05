"""
from src.utils.detector_colunas import extrair_numeros_sorteio
Atualizador de Abas de Universos Reduzidos - MegaCLI v6.0

Cria abas TOP_20N e TOP_9N no Excel com validação histórica
dos últimos 100 jogos para cada universo.

Autor: MegaCLI Team
Data: 22/01/2026
Versão: 1.0.0
"""

import pandas as pd
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime


def criar_aba_universo(
    arquivo_excel: str,
    nome_aba: str,
    numeros_universo: List[int],
    df_historico: pd.DataFrame,
    janela_validacao: int = 100,
    verbose: bool = True
) -> pd.DataFrame:
    """
    Cria aba de validação para universo reduzido (20 ou 9 números).
    
    Args:
        arquivo_excel: Caminho do Excel
        nome_aba: Nome da aba (TOP_20N ou TOP_9N)
        numeros_universo: Lista de números do universo
        df_historico: DataFrame com histórico
        janela_validacao: Janela de validação
        verbose: Se True, exibe informações
        
    Returns:
        DataFrame com validação
    """
    if verbose:
        print(f"\n📊 Criando aba {nome_aba}...")
    
    # Pegar últimos N sorteios
    df_ultimos = df_historico.tail(janela_validacao).copy()
    
    # Criar DataFrame de validação
    validacao = []
    
    for idx, row in df_ultimos.iterrows():
        concurso = row['Concurso']
        numeros_sorteio = extrair_numeros_sorteio(row)
        
        # Contar acertos
        acertos = len(set(numeros_sorteio) & set(numeros_universo))
        
        # Verificar se todos os números estão no universo
        todos_no_universo = all(n in numeros_universo for n in numeros_sorteio)
        
        validacao.append({
            'Concurso': concurso,
            'Num1': numeros_sorteio[0],
            'Num2': numeros_sorteio[1],
            'Num3': numeros_sorteio[2],
            'Num4': numeros_sorteio[3],
            'Num5': numeros_sorteio[4],
            'Num6': numeros_sorteio[5],
            'Nums_Universo': acertos,
            'Cobertura_Total': '✅ SIM' if todos_no_universo else '❌ NÃO',
            'Status': 'COBERTO' if todos_no_universo else ('PARCIAL' if acertos >= 4 else 'BAIXO')
        })
    
    df_validacao = pd.DataFrame(validacao)
    
    # Calcular estatísticas
    total_cobertos = (df_validacao['Nums_Universo'] == 6).sum()
    total_5_plus = (df_validacao['Nums_Universo'] >= 5).sum()
    total_4_plus = (df_validacao['Nums_Universo'] >= 4).sum()
    
    taxa_cobertura = (total_cobertos / janela_validacao) * 100
    
    # Adicionar ao Excel
    wb = load_workbook(arquivo_excel)
    
    # Remover aba se já existir
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]
    
    # Criar nova aba
    ws = wb.create_sheet(nome_aba)
    
    # Cabeçalho
    ws['A1'] = f'VALIDAÇÃO HISTÓRICA - {nome_aba}'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws.merge_cells('A1:J1')
    
    # Informações do universo
    row_atual = 2
    ws[f'A{row_atual}'] = f'Universo ({len(numeros_universo)} números):'
    ws[f'A{row_atual}'].font = Font(bold=True)
    ws[f'B{row_atual}'] = '-'.join(f"{n:02d}" for n in numeros_universo)
    row_atual += 1
    
    ws[f'A{row_atual}'] = f'Janela de Validação:'
    ws[f'A{row_atual}'].font = Font(bold=True)
    ws[f'B{row_atual}'] = f'{janela_validacao} sorteios'
    row_atual += 1
    
    ws[f'A{row_atual}'] = f'Data de Geração:'
    ws[f'A{row_atual}'].font = Font(bold=True)
    ws[f'B{row_atual}'] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    row_atual += 2
    
    # Estatísticas
    ws[f'A{row_atual}'] = 'ESTATÍSTICAS'
    ws[f'A{row_atual}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'A{row_atual}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    ws.merge_cells(f'A{row_atual}:D{row_atual}')
    row_atual += 1
    
    ws[f'A{row_atual}'] = 'Cobertura Total (6 números):'
    ws[f'B{row_atual}'] = f'{total_cobertos} ({taxa_cobertura:.1f}%)'
    row_atual += 1
    
    ws[f'A{row_atual}'] = 'Cobertura 5+ números:'
    ws[f'B{row_atual}'] = f'{total_5_plus} ({(total_5_plus/janela_validacao)*100:.1f}%)'
    row_atual += 1
    
    ws[f'A{row_atual}'] = 'Cobertura 4+ números:'
    ws[f'B{row_atual}'] = f'{total_4_plus} ({(total_4_plus/janela_validacao)*100:.1f}%)'
    row_atual += 2
    
    # Dados de validação
    ws[f'A{row_atual}'] = 'VALIDAÇÃO DETALHADA'
    ws[f'A{row_atual}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws[f'A{row_atual}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    ws.merge_cells(f'A{row_atual}:J{row_atual}')
    row_atual += 1
    
    # Adicionar DataFrame
    for r_idx, row in enumerate(dataframe_to_rows(df_validacao, index=False, header=True), row_atual):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Formatar cabeçalho
            if r_idx == row_atual:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Colorir linhas com cobertura total
            elif c_idx == 9 and value == '✅ SIM':  # Coluna Cobertura_Total
                for col in range(1, 11):
                    ws.cell(row=r_idx, column=col).fill = PatternFill(
                        start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'
                    )
    
    # Auto-ajustar colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar
    wb.save(arquivo_excel)
    
    if verbose:
        print(f"✅ Aba {nome_aba} criada com sucesso!")
        print(f"   • Sorteios analisados: {janela_validacao}")
        print(f"   • Cobertura total: {total_cobertos} ({taxa_cobertura:.1f}%)")
    
    return df_validacao


def atualizar_abas_universos(
    arquivo_excel: str,
    numeros_20: List[int],
    numeros_9: List[int],
    df_historico: pd.DataFrame,
    janela_validacao: int = 100,
    verbose: bool = True
) -> None:
    """
    Atualiza ambas as abas TOP_20N e TOP_9N no Excel.
    
    Args:
        arquivo_excel: Caminho do Excel
        numeros_20: Lista com 20 números
        numeros_9: Lista com 9 números
        df_historico: DataFrame com histórico
        janela_validacao: Janela de validação
        verbose: Se True, exibe informações
    """
    if verbose:
        print(f"\n💾 Atualizando abas de universos reduzidos...")
    
    # Criar aba TOP_20N
    df_val_20 = criar_aba_universo(
        arquivo_excel,
        'TOP_20N',
        numeros_20,
        df_historico,
        janela_validacao,
        verbose
    )
    
    # Criar aba TOP_9N
    df_val_9 = criar_aba_universo(
        arquivo_excel,
        'TOP_9N',
        numeros_9,
        df_historico,
        janela_validacao,
        verbose
    )
    
    if verbose:
        print(f"\n✅ Abas atualizadas com sucesso!")
        print(f"   • TOP_20N: {len(df_val_20)} sorteios")
        print(f"   • TOP_9N: {len(df_val_9)} sorteios")


# Exports
__all__ = [
    'criar_aba_universo',
    'atualizar_abas_universos'
]


# Teste standalone
if __name__ == "__main__":
    print("\n🧪 Testando Atualizador de Abas...\n")
    
    import sys
    from pathlib import Path
    PROJECT_ROOT = Path(__file__).parent.parent.parent
    sys.path.insert(0, str(PROJECT_ROOT))
    
    from src.core.config import ARQUIVO_HISTORICO, RESULTADO_DIR
    
    # Carregar histórico
    df_historico = pd.read_excel(str(ARQUIVO_HISTORICO), sheet_name='MEGA SENA')
    print(f"✅ {len(df_historico)} sorteios carregados")
    
    # Números de teste
    numeros_20_teste = [5, 10, 12, 15, 18, 20, 23, 27, 30, 34, 37, 38, 40, 42, 45, 48, 50, 53, 56, 58]
    numeros_9_teste = [5, 10, 23, 27, 34, 38, 42, 48, 53]
    
    # Atualizar abas
    arquivo_excel = RESULTADO_DIR / 'ANALISE_HISTORICO_COMPLETO.xlsx'
    
    atualizar_abas_universos(
        str(arquivo_excel),
        numeros_20_teste,
        numeros_9_teste,
        df_historico,
        janela_validacao=100,
        verbose=True
    )
    
    print("\n✅ Módulo funcionando corretamente!\n")
