"""
Validador Retroativo Expandido - MegaCLI v6.1

Sistema completo de validação retroativa com:
- Validação multi-nível (TOP 30, 20, 10, 9)
- Consulta à IA para análise de indicadores
- Reavaliação probabilística
- Análise de grupos de indicadores ótimos
- Auto-aprendizado contínuo

Autor: MegaCLI Team
Data: 01/02/2026
Versão: 2.0.0
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional
import json
from collections import Counter
from itertools import combinations

# Importações do projeto
from src.core.config import ARQUIVO_HISTORICO, RESULTADO_DIR, DADOS_DIR
from src.utils.detector_colunas import extrair_numeros_sorteio
from src.core.conexao_ia import conectar_ia
from src.core.previsao_30n import selecionar_top_30_numeros, refinar_selecao


# ============================================================================
# FASE 1: VALIDAÇÃO MULTI-NÍVEL
# ============================================================================

def validar_multi_nivel(
    numeros_sorteados: List[int],
    df_historico: pd.DataFrame,
    ranking_indicadores: List[Dict]
) -> Dict[str, Any]:
    """
    Gera previsões em 4 níveis (TOP 30, 20, 10, 9) e valida acertos.
    
    Args:
        numeros_sorteados: Números do sorteio real
        df_historico: DataFrame com histórico até este sorteio
        ranking_indicadores: Lista com ranking de indicadores
        
    Returns:
        Dict com resultados de validação em todos os níveis
    """
    print("   📊 Gerando previsões em múltiplos níveis...")
    
    # Gerar TOP 30
    top_30, scores_30, rastro_votos, detalhe_numeros = selecionar_top_30_numeros(
        df_historico,
        ranking_indicadores,
        verbose=False
    )
    
    # Refinar para obter sequência ordenada
    lista_refinada, _ = refinar_selecao(top_30, scores_30, df_historico, verbose=False)
    
    # Definir subconjuntos
    sets_numeros = {
        30: set(lista_refinada[:30]),
        20: set(lista_refinada[:20]),
        10: set(lista_refinada[:10]),
        9: set(lista_refinada[:9])
    }
    
    # Calcular acertos para cada nível
    numeros_sorteados_set = set(numeros_sorteados)
    resultados = {}
    
    for nivel, numeros_previstos in sets_numeros.items():
        acertos_set = numeros_sorteados_set & numeros_previstos
        perdidos_set = numeros_sorteados_set - numeros_previstos
        
        resultados[f'top_{nivel}'] = {
            'numeros': sorted(list(numeros_previstos)),
            'acertos': len(acertos_set),
            'numeros_acertados': sorted(list(acertos_set)),
            'numeros_perdidos': sorted(list(perdidos_set)),
            'taxa_acerto': len(acertos_set) / 6.0
        }
    
    # Adicionar detalhes de rastreamento
    resultados['rastro_votos'] = rastro_votos
    resultados['detalhe_numeros'] = detalhe_numeros
    
    return resultados


# ============================================================================
# FASE 3: CONSULTA À IA PARA ANÁLISE RETROATIVA
# ============================================================================

def consultar_ia_analise_retroativa(
    concurso: int,
    numeros_sorteados: List[int],
    resultado_validacao: Dict[str, Any],
    ranking_usado: List[Dict]
) -> Optional[Dict[str, Any]]:
    """
    Consulta IA para identificar quais indicadores deveriam ter sido priorizados.
    
    Args:
        concurso: Número do concurso
        numeros_sorteados: Números realmente sorteados
        resultado_validacao: Resultado da validação multi-nível
        ranking_usado: Ranking de indicadores que foi usado
        
    Returns:
        Dict com sugestões da IA ou None se falhar
    """
    print("   🤖 Consultando IA para análise retroativa...")
    
    # Formatação dos números
    nums_str = "-".join(f"{n:02d}" for n in sorted(numeros_sorteados))
    
    # TOP 30 usado
    top_30_info = resultado_validacao['top_30']
    top_30_str = "-".join(f"{n:02d}" for n in top_30_info['numeros'][:30])
    acertos_30 = top_30_info['acertos']
    perdidos = top_30_info['numeros_perdidos']
    
    # Top 5 indicadores usados
    top_indicadores = sorted(ranking_usado, key=lambda x: x.get('relevancia', 0), reverse=True)[:10]
    indicadores_str = ", ".join([f"{ind['indicador']} ({ind.get('relevancia', 0):.1f})" for ind in top_indicadores])
    
    # Montar prompt
    prompt = f"""Você é um especialista em análise estatística de loterias.

RESULTADO REAL - Concurso {concurso}:
Números sorteados: {nums_str}

PREVISÃO GERADA (com indicadores atuais):
TOP 30: {top_30_str}
Acertos: {acertos_30}/6 números ({acertos_30/6*100:.1f}%)
Números perdidos: {perdidos}

INDICADORES USADOS (Top 10 com maior peso):
{indicadores_str}

ANÁLISE SOLICITADA:
1. Quais indicadores deveriam ter peso MAIOR para capturar os números perdidos?
2. Quais indicadores deveriam ter peso MENOR (poluíram a previsão)?
3. Qual combinação de indicadores teria melhor desempenho para este padrão?

RESPONDA EM JSON com a seguinte estrutura:
{{
  "indicadores_aumentar": [
    {{"nome": "NomeIndicador", "peso_sugerido": 90, "motivo": "explicação curta"}}
  ],
  "indicadores_diminuir": [
    {{"nome": "NomeIndicador", "peso_sugerido": 40, "motivo": "explicação curta"}}
  ],
  "grupo_otimo": ["Ind1", "Ind2", "Ind3", "Ind4", "Ind5"],
  "analise_geral": "Análise resumida do padrão deste sorteio",
  "confianca": 0.85
}}

IMPORTANTE: Baseie-se nos números perdidos {perdidos} e sugira indicadores que teriam capturado esses números.
"""
    
    try:
        llm = conectar_ia(temperatura=0.2, verbose=False)
        if not llm:
            print("   ⚠️  Falha ao conectar com IA")
            return None
        
        resposta = llm.invoke(prompt)
        
        # Parse JSON
        import re
        content = resposta.content
        
        # Tentar extrair JSON
        json_match = re.search(r'```json\n(.*?)\n```', content, re.DOTALL)
        if json_match:
            json_str = json_match.group(1)
        else:
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            if json_match:
                json_str = json_match.group(0)
            else:
                print("   ⚠️  Não foi possível extrair JSON da resposta da IA")
                return None
        
        sugestoes = json.loads(json_str)
        print(f"   ✅ IA sugeriu ajustes em {len(sugestoes.get('indicadores_aumentar', []))} indicadores")
        
        return sugestoes
        
    except Exception as e:
        print(f"   ⚠️  Erro ao consultar IA: {e}")
        return None


# ============================================================================
# FASE 4: REAVALIAÇÃO PROBABILÍSTICA
# ============================================================================

def reavaliar_com_ia(
    numeros_sorteados: List[int],
    df_historico: pd.DataFrame,
    ranking_original: List[Dict],
    sugestoes_ia: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Testa se com os pesos sugeridos pela IA teríamos acertado mais.
    
    Args:
        numeros_sorteados: Números realmente sorteados
        df_historico: Histórico até este sorteio
        ranking_original: Ranking usado originalmente
        sugestoes_ia: Sugestões da IA
        
    Returns:
        Dict com comparativo antes/depois
    """
    print("   🔄 Reavaliando com pesos sugeridos pela IA...")
    
    # Criar ranking ajustado
    ranking_ajustado = ranking_original.copy()
    
    # Aplicar sugestões de aumento
    for sug in sugestoes_ia.get('indicadores_aumentar', []):
        for ind in ranking_ajustado:
            if ind['indicador'].lower() == sug['nome'].lower():
                ind['relevancia'] = sug['peso_sugerido']
                break
    
    # Aplicar sugestões de diminuição
    for sug in sugestoes_ia.get('indicadores_diminuir', []):
        for ind in ranking_ajustado:
            if ind['indicador'].lower() == sug['nome'].lower():
                ind['relevancia'] = sug['peso_sugerido']
                break
    
    # Gerar nova previsão com ranking ajustado
    resultado_ia = validar_multi_nivel(numeros_sorteados, df_historico, ranking_ajustado)
    
    acertos_ia_30 = resultado_ia['top_30']['acertos']
    acertos_ia_9 = resultado_ia['top_9']['acertos']
    
    print(f"   📈 Com IA: {acertos_ia_30}/6 no TOP 30, {acertos_ia_9}/6 no TOP 9")
    
    return {
        'ranking_ajustado': ranking_ajustado,
        'resultado_ia': resultado_ia,
        'acertos_top_30_ia': acertos_ia_30,
        'acertos_top_9_ia': acertos_ia_9
    }


# ============================================================================
# FASE 2: ATUALIZAÇÃO DO EXCEL
# ============================================================================

def atualizar_excel_validacao(
    resultados_validacao: List[Dict[str, Any]],
    arquivo_excel: Path = None
) -> None:
    """
    Atualiza aba VALIDACAO_RETROATIVA no Excel com resultados.
    
    Args:
        resultados_validacao: Lista de dicionários com resultados
        arquivo_excel: Path para o Excel (default: ANALISE_HISTORICO_COMPLETO.xlsx)
    """
    if arquivo_excel is None:
        arquivo_excel = RESULTADO_DIR / 'ANALISE_HISTORICO_COMPLETO.xlsx'
    
    print(f"\n💾 Atualizando {arquivo_excel.name}...")
    
    # Preparar dados para DataFrame
    registros = []
    for res in resultados_validacao:
        registro = {
            'Concurso': res['concurso'],
            'Data': res['data'],
            'Numeros_Sorteados': "-".join(f"{n:02d}" for n in sorted(res['numeros_sorteados'])),
            
            # TOP 30
            'TOP_30': "-".join(f"{n:02d}" for n in res['validacao']['top_30']['numeros'][:30]),
            'Acertos_30': res['validacao']['top_30']['acertos'],
            'Taxa_30': f"{res['validacao']['top_30']['taxa_acerto']*100:.1f}%",
            
            # TOP 20
            'TOP_20': "-".join(f"{n:02d}" for n in res['validacao']['top_20']['numeros'][:20]),
            'Acertos_20': res['validacao']['top_20']['acertos'],
            
            # TOP 10
            'TOP_10': "-".join(f"{n:02d}" for n in res['validacao']['top_10']['numeros'][:10]),
            'Acertos_10': res['validacao']['top_10']['acertos'],
            
            # TOP 9
            'TOP_9': "-".join(f"{n:02d}" for n in res['validacao']['top_9']['numeros'][:9]),
            'Acertos_9': res['validacao']['top_9']['acertos'],
            
            # IA
            'Usou_IA': 'Sim' if res.get('ia_sugestoes') else 'Não',
            'Acertos_IA_30': res.get('reavaliacao', {}).get('acertos_top_30_ia', '-'),
            'Melhoria': res.get('reavaliacao', {}).get('acertos_top_30_ia', 0) - res['validacao']['top_30']['acertos'] if res.get('reavaliacao') else 0
        }
        registros.append(registro)
    
    df_novo = pd.DataFrame(registros)
    
    # Salvar no Excel
    try:
        from openpyxl import load_workbook
        
        if arquivo_excel.exists():
            wb = load_workbook(arquivo_excel)
        else:
            from openpyxl import Workbook
            wb = Workbook()
        
        # Remover aba se existir
        if 'VALIDACAO_RETROATIVA' in wb.sheetnames:
            del wb['VALIDACAO_RETROATIVA']
        
        # Criar nova aba
        ws = wb.create_sheet('VALIDACAO_RETROATIVA')
        
        # Escrever dados
        from openpyxl.utils.dataframe import dataframe_to_rows
        for row in dataframe_to_rows(df_novo, index=False, header=True):
            ws.append(row)
        
        wb.save(arquivo_excel)
        print(f"   ✅ Aba VALIDACAO_RETROATIVA atualizada com {len(registros)} registros")
        
    except Exception as e:
        print(f"   ⚠️  Erro ao atualizar Excel: {e}")


# ============================================================================
# EM DESENVOLVIMENTO: FASE 5 - ANÁLISE DE GRUPOS
# (Será implementada no próximo arquivo)
# ============================================================================

