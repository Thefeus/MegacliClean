"""
Validador Histórico para Aba Excel - MegaCLI v6.0

Cria aba de validação histórica com últimos 100 jogos e previsão
para próximo sorteio.

Autor: MegaCLI Team
Data: 22/01/2026
Versão: 1.0.0
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime


def validar_jogos_historico_100(
    jogos_gerados: List[Dict],
    df_historico: pd.DataFrame,
    universo_reduzido: List[int] = None,
    verbose: bool = True
) -> pd.DataFrame:
    """
    Valida jogos gerados contra últimos 100 sorteios históricos.
    
    Args:
        jogos_gerados: Lista de jogos gerados
        df_historico: DataFrame com histórico completo
        universo_reduzido: Lista com 20 números do universo (opcional)
        verbose: Se True, exibe informações
        
    Returns:
        DataFrame com validação completa
    """
    if verbose:
        print(f"\n📊 Validando jogos contra últimos 100 sorteios...")
    
    # Pegar últimos 100 sorteios
    df_ultimos_100 = df_historico.tail(100).copy()
    
    resultados = []
    
    for jogo in jogos_gerados:
        numeros_jogo = set(jogo['numeros'])
        
        # Contar acertos em cada sorteio
        acertos_por_sorteio = []
        
        for _, sorteio in df_ultimos_100.iterrows():
            numeros_sorteio = set([sorteio[f'Bola{i}'] for i in range(1, 7)])
            acertos = len(numeros_jogo & numeros_sorteio)
            acertos_por_sorteio.append(acertos)
        
        # Estatísticas
        total_acertos_3 = sum(1 for a in acertos_por_sorteio if a >= 3)
        total_acertos_4 = sum(1 for a in acertos_por_sorteio if a >= 4)
        total_acertos_5 = sum(1 for a in acertos_por_sorteio if a >= 5)
        total_acertos_6 = sum(1 for a in acertos_por_sorteio if a == 6)
        
        # Verificar se números estão no universo reduzido
        if universo_reduzido:
            numeros_no_universo = len(numeros_jogo & set(universo_reduzido))
        else:
            numeros_no_universo = 6  # Todos
        
        resultado = {
            'Rank': jogo['rank'],
            'Jogo': '-'.join(f"{n:02d}" for n in jogo['numeros']),
            'Num1': jogo['numeros'][0],
            'Num2': jogo['numeros'][1],
            'Num3': jogo['numeros'][2],
            'Num4': jogo['numeros'][3],
            'Num5': jogo['numeros'][4],
            'Num6': jogo['numeros'][5],
            'Score': jogo['score'],
            'Nums_Universo': numeros_no_universo,
            'Acertos_3+': total_acertos_3,
            'Acertos_4+': total_acertos_4,
            'Acertos_5+': total_acertos_5,
            'Acertos_6': total_acertos_6,
            'Taxa_3+_%': (total_acertos_3 / 100) * 100,
            'Taxa_4+_%': (total_acertos_4 / 100) * 100,
            'Taxa_5+_%': (total_acertos_5 / 100) * 100,
            'Taxa_6_%': (total_acertos_6 / 100) * 100,
            'Melhor_Acerto': max(acertos_por_sorteio),
            'Media_Acertos': np.mean(acertos_por_sorteio)
        }
        
        resultados.append(resultado)
    
    df_validacao = pd.DataFrame(resultados)
    
    if verbose:
        print(f"✅ Validação concluída para {len(jogos_gerados)} jogos")
        print(f"   Média de acertos 3+: {df_validacao['Acertos_3+'].mean():.2f}")
        print(f"   Média de acertos 4+: {df_validacao['Acertos_4+'].mean():.2f}")
    
    return df_validacao


def criar_previsao_proximo_sorteio(
    jogos_gerados: List[Dict],
    df_historico: pd.DataFrame,
    universo_reduzido: List[int] = None
) -> pd.DataFrame:
    """
    Cria DataFrame com previsão para próximo sorteio.
    
    Args:
        jogos_gerados: Lista de jogos gerados
        df_historico: DataFrame com histórico
        universo_reduzido: Lista com 20 números (opcional)
        
    Returns:
        DataFrame com previsão
    """
    # Próximo concurso
    ultimo_concurso = df_historico.iloc[-1]['Concurso']
    proximo_concurso = ultimo_concurso + 1
    
    previsoes = []
    
    for jogo in jogos_gerados:
        numeros_jogo = set(jogo['numeros'])
        
        # Verificar números no universo
        if universo_reduzido:
            numeros_no_universo = len(numeros_jogo & set(universo_reduzido))
        else:
            numeros_no_universo = 6
        
        previsao = {
            'Concurso': proximo_concurso,
            'Rank': jogo['rank'],
            'Jogo': '-'.join(f"{n:02d}" for n in jogo['numeros']),
            'Num1': jogo['numeros'][0],
            'Num2': jogo['numeros'][1],
            'Num3': jogo['numeros'][2],
            'Num4': jogo['numeros'][3],
            'Num5': jogo['numeros'][4],
            'Num6': jogo['numeros'][5],
            'Score': jogo['score'],
            'Probabilidade_%': jogo.get('probabilidade', jogo['score']),
            'Nums_Universo': numeros_no_universo,
            'Confianca': jogo.get('confianca', 'MÉDIA'),
            'Acertos': 0,  # Será preenchido após sorteio
            'Status': 'PENDENTE'
        }
        
        previsoes.append(previsao)
    
    return pd.DataFrame(previsoes)


def adicionar_aba_validacao_excel(
    arquivo_excel: str,
    df_validacao: pd.DataFrame,
    df_previsao: pd.DataFrame,
    universo_reduzido: List[int] = None,
    verbose: bool = True
) -> None:
    """
    Adiciona aba de validação histórica ao Excel.
    
    Args:
        arquivo_excel: Caminho do arquivo Excel
        df_validacao: DataFrame com validação dos 100 jogos
        df_previsao: DataFrame com previsão para próximo sorteio
        universo_reduzido: Lista com 20 números (opcional)
        verbose: Se True, exibe informações
    """
    if verbose:
        print(f"\n💾 Adicionando aba de validação ao Excel...")
    
    # Abrir workbook
    wb = load_workbook(arquivo_excel)
    
    # Remover aba se já existir
    if 'VALIDAÇÃO_100' in wb.sheetnames:
        del wb['VALIDAÇÃO_100']
    
    # Criar nova aba
    ws = wb.create_sheet('VALIDAÇÃO_100')
    
    # Adicionar cabeçalho
    ws['A1'] = 'VALIDAÇÃO HISTÓRICA - ÚLTIMOS 100 JOGOS'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws.merge_cells('A1:P1')
    
    # Informações do universo
    row_atual = 2
    if universo_reduzido:
        ws[f'A{row_atual}'] = 'Universo Reduzido (20 números):'
        ws[f'A{row_atual}'].font = Font(bold=True)
        ws[f'B{row_atual}'] = '-'.join(f"{n:02d}" for n in universo_reduzido)
        row_atual += 1
    
    ws[f'A{row_atual}'] = f'Data de Geração: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
    row_atual += 2
    
    # Adicionar dados de validação
    ws[f'A{row_atual}'] = 'VALIDAÇÃO HISTÓRICA (100 SORTEIOS)'
    ws[f'A{row_atual}'].font = Font(bold=True, size=12)
    ws[f'A{row_atual}'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
    ws[f'A{row_atual}'].font = Font(bold=True, size=12, color='FFFFFF')
    ws.merge_cells(f'A{row_atual}:P{row_atual}')
    row_atual += 1
    
    # Adicionar DataFrame de validação
    for r_idx, row in enumerate(dataframe_to_rows(df_validacao, index=False, header=True), row_atual):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Formatar cabeçalho
            if r_idx == row_atual:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
    
    row_atual += len(df_validacao) + 3
    
    # Adicionar previsão para próximo sorteio
    ws[f'A{row_atual}'] = 'PREVISÃO PARA PRÓXIMO SORTEIO'
    ws[f'A{row_atual}'].font = Font(bold=True, size=12)
    ws[f'A{row_atual}'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    ws[f'A{row_atual}'].font = Font(bold=True, size=12, color='000000')
    ws.merge_cells(f'A{row_atual}:N{row_atual}')
    row_atual += 1
    
    # Adicionar DataFrame de previsão
    for r_idx, row in enumerate(dataframe_to_rows(df_previsao, index=False, header=True), row_atual):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Formatar cabeçalho
            if r_idx == row_atual:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-ajustar colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar
    wb.save(arquivo_excel)
    
    if verbose:
        print(f"✅ Aba 'VALIDAÇÃO_100' adicionada com sucesso!")
        print(f"   • Validação histórica: {len(df_validacao)} jogos")
        print(f"   • Previsão próximo sorteio: {len(df_previsao)} jogos")


# Exports
__all__ = [
    'validar_jogos_historico_100',
    'criar_previsao_proximo_sorteio',
    'adicionar_aba_validacao_excel'
]


# Teste standalone
if __name__ == "__main__":
    print("\n🧪 Testando Validador Histórico...\n")
    
    import sys
    from pathlib import Path
    PROJECT_ROOT = Path(__file__).parent.parent.parent
    sys.path.insert(0, str(PROJECT_ROOT))
    
    from src.core.config import ARQUIVO_HISTORICO, RESULTADO_DIR
    
    # Carregar histórico
    df_historico = pd.read_excel(str(ARQUIVO_HISTORICO), sheet_name='MEGA SENA')
    print(f"✅ {len(df_historico)} sorteios carregados")
    
    # Jogos de teste
    jogos_teste = [
        {
            'rank': i,
            'numeros': [5, 10, 20, 30, 40, 50],
            'score': 85.0,
            'probabilidade': 85.0,
            'confianca': 'ALTA ⭐⭐⭐'
        }
        for i in range(1, 11)
    ]
    
    # Universo de teste
    universo_teste = [5, 10, 12, 15, 18, 20, 23, 27, 30, 34, 37, 38, 40, 42, 45, 48, 50, 53, 56, 58]
    
    # Validar
    df_validacao = validar_jogos_historico_100(
        jogos_teste,
        df_historico,
        universo_reduzido=universo_teste,
        verbose=True
    )
    
    # Criar previsão
    df_previsao = criar_previsao_proximo_sorteio(
        jogos_teste,
        df_historico,
        universo_reduzido=universo_teste
    )
    
    print(f"\n✅ Validação criada: {len(df_validacao)} linhas")
    print(f"✅ Previsão criada: {len(df_previsao)} linhas")
    
    print("\n✅ Módulo funcionando corretamente!\n")
