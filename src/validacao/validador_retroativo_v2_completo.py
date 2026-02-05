"""
Validador Retroativo Expandido - MegaCLI v6.1

Sistema completo de validação retroativa com:
- Validação multi-nível (TOP 30, 20, 10, 9)
- Consulta à IA para análise de indicadores
- Reavaliação probabilística
- Análise de grupos de indicadores ótimos
- Auto-aprendizado contínuo

Autor: MegaCLI Team
Data: 01/02/2026
Versão: 2.0.0
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Any, Optional
import json
from collections import Counter
from itertools import combinations

# Importações do projeto
from src.core.config import ARQUIVO_HISTORICO, RESULTADO_DIR, DADOS_DIR
from src.utils.detector_colunas import extrair_numeros_sorteio
from src.core.conexao_ia import conectar_ia
from src.core.previsao_30n import selecionar_top_30_numeros, refinar_selecao


# ============================================================================
# FASE 1: VALIDAÇÃO MULTI-NÍVEL
# ============================================================================

def validar_multi_nivel(
    numeros_sorteados: List[int],
    df_historico: pd.DataFrame,
    ranking_indicadores: List[Dict]
) -> Dict[str, Any]:
    """
    Gera previsões em 4 níveis (TOP 30, 20, 10, 9) e valida acertos.
    
    Args:
        numeros_sorteados: Números do sorteio real
        df_historico: DataFrame com histórico até este sorteio
        ranking_indicadores: Lista com ranking de indicadores
        
    Returns:
        Dict com resultados de validação em todos os níveis
    """
    print("   📊 Gerando previsões em múltiplos níveis...")
    
    # Gerar TOP 30
    top_30, scores_30, rastro_votos, detalhe_numeros = selecionar_top_30_numeros(
        df_historico,
        ranking_indicadores,
        verbose=False
    )
    
    # Refinar para obter sequência ordenada
    lista_refinada, _ = refinar_selecao(top_30, scores_30, df_historico, verbose=False)
    
    # Definir subconjuntos
    sets_numeros = {
        30: set(lista_refinada[:30]),
        20: set(lista_refinada[:20]),
        10: set(lista_refinada[:10]),
        9: set(lista_refinada[:9])
    }
    
    # Calcular acertos para cada nível
    numeros_sorteados_set = set(numeros_sorteados)
    resultados = {}
    
    for nivel, numeros_previstos in sets_numeros.items():
        acertos_set = numeros_sorteados_set & numeros_previstos
        perdidos_set = numeros_sorteados_set - numeros_previstos
        
        resultados[f'top_{nivel}'] = {
            'numeros': sorted(list(numeros_previstos)),
            'acertos': len(acertos_set),
            'numeros_acertados': sorted(list(acertos_set)),
            'numeros_perdidos': sorted(list(perdidos_set)),
            'taxa_acerto': len(acertos_set) / 6.0
        }
    
    # Adicionar detalhes de rastreamento
    resultados['rastro_votos'] = rastro_votos
    resultados['detalhe_numeros'] = detalhe_numeros
    
    return resultados


# ============================================================================
# FASE 3: CONSULTA À IA PARA ANÁLISE RETROATIVA
# ============================================================================

def consultar_ia_analise_retroativa(
    concurso: int,
    numeros_sorteados: List[int],
    resultado_validacao: Dict[str, Any],
    ranking_usado: List[Dict]
) -> Optional[Dict[str, Any]]:
    """
    Consulta IA para identificar quais indicadores deveriam ter sido priorizados.
    
    Args:
        concurso: Número do concurso
        numeros_sorteados: Números realmente sorteados
        resultado_validacao: Resultado da validação multi-nível
        ranking_usado: Ranking de indicadores que foi usado
        
    Returns:
        Dict com sugestões da IA ou None se falhar
    """
    print("   🤖 Consultando IA para análise retroativa...")
    
    # Formatação dos números
    nums_str = "-".join(f"{n:02d}" for n in sorted(numeros_sorteados))
    
    # TOP 30 usado
    top_30_info = resultado_validacao['top_30']
    top_30_str = "-".join(f"{n:02d}" for n in top_30_info['numeros'][:30])
    acertos_30 = top_30_info['acertos']
    perdidos = top_30_info['numeros_perdidos']
    
    # Top 5 indicadores usados
    top_indicadores = sorted(ranking_usado, key=lambda x: x.get('relevancia', 0), reverse=True)[:10]
    indicadores_str = ", ".join([f"{ind['indicador']} ({ind.get('relevancia', 0):.1f})" for ind in top_indicadores])
    
    # Montar prompt
    prompt = f"""Você é um especialista em análise estatística de loterias.

RESULTADO REAL - Concurso {concurso}:
Números sorteados: {nums_str}

PREVISÃO GERADA (com indicadores atuais):
TOP 30: {top_30_str}
Acertos: {acertos_30}/6 números ({acertos_30/6*100:.1f}%)
Números perdidos: {perdidos}

INDICADORES USADOS (Top 10 com maior peso):
{indicadores_str}

ANÁLISE SOLICITADA:
1. Quais indicadores deveriam ter peso MAIOR para capturar os números perdidos?
2. Quais indicadores deveriam ter peso MENOR (poluíram a previsão)?
3. Qual combinação de indicadores teria melhor desempenho para este padrão?

RESPONDA EM JSON com a seguinte estrutura:
{{
  "indicadores_aumentar": [
    {{"nome": "NomeIndicador", "peso_sugerido": 90, "motivo": "explicação curta"}}
  ],
  "indicadores_diminuir": [
    {{"nome": "NomeIndicador", "peso_sugerido": 40, "motivo": "explicação curta"}}
  ],
  "grupo_otimo": ["Ind1", "Ind2", "Ind3", "Ind4", "Ind5"],
  "analise_geral": "Análise resumida do padrão deste sorteio",
  "confianca": 0.85
}}

IMPORTANTE: Baseie-se nos números perdidos {perdidos} e sugira indicadores que teriam capturado esses números.
"""
    
    try:
        llm = conectar_ia(temperatura=0.2, verbose=False)
        if not llm:
            print("   ⚠️  Falha ao conectar com IA")
            return None
        
        resposta = llm.invoke(prompt)
        
        # Parse JSON
        import re
        content = resposta.content
        
        # Tentar extrair JSON
        json_match = re.search(r'```json\n(.*?)\n```', content, re.DOTALL)
        if json_match:
            json_str = json_match.group(1)
        else:
            json_match = re.search(r'\{.*\}', content, re.DOTALL)
            if json_match:
                json_str = json_match.group(0)
            else:
                print("   ⚠️  Não foi possível extrair JSON da resposta da IA")
                return None
        
        sugestoes = json.loads(json_str)
        print(f"   ✅ IA sugeriu ajustes em {len(sugestoes.get('indicadores_aumentar', []))} indicadores")
        
        return sugestoes
        
    except Exception as e:
        print(f"   ⚠️  Erro ao consultar IA: {e}")
        return None


# ============================================================================
# FASE 4: REAVALIAÇÃO PROBABILÍSTICA
# ============================================================================

def reavaliar_com_ia(
    numeros_sorteados: List[int],
    df_historico: pd.DataFrame,
    ranking_original: List[Dict],
    sugestoes_ia: Dict[str, Any]
) -> Dict[str, Any]:
    """
    Testa se com os pesos sugeridos pela IA teríamos acertado mais.
    
    Args:
        numeros_sorteados: Números realmente sorteados
        df_historico: Histórico até este sorteio
        ranking_original: Ranking usado originalmente
        sugestoes_ia: Sugestões da IA
        
    Returns:
        Dict com comparativo antes/depois
    """
    print("   🔄 Reavaliando com pesos sugeridos pela IA...")
    
    # Criar ranking ajustado
    ranking_ajustado = ranking_original.copy()
    
    # Aplicar sugestões de aumento
    for sug in sugestoes_ia.get('indicadores_aumentar', []):
        for ind in ranking_ajustado:
            if ind['indicador'].lower() == sug['nome'].lower():
                ind['relevancia'] = sug['peso_sugerido']
                break
    
    # Aplicar sugestões de diminuição
    for sug in sugestoes_ia.get('indicadores_diminuir', []):
        for ind in ranking_ajustado:
            if ind['indicador'].lower() == sug['nome'].lower():
                ind['relevancia'] = sug['peso_sugerido']
                break
    
    # Gerar nova previsão com ranking ajustado
    resultado_ia = validar_multi_nivel(numeros_sorteados, df_historico, ranking_ajustado)
    
    acertos_ia_30 = resultado_ia['top_30']['acertos']
    acertos_ia_9 = resultado_ia['top_9']['acertos']
    
    print(f"   📈 Com IA: {acertos_ia_30}/6 no TOP 30, {acertos_ia_9}/6 no TOP 9")
    
    return {
        'ranking_ajustado': ranking_ajustado,
        'resultado_ia': resultado_ia,
        'acertos_top_30_ia': acertos_ia_30,
        'acertos_top_9_ia': acertos_ia_9
    }


# ============================================================================
# FASE 2: ATUALIZAÇÃO DO EXCEL
# ============================================================================

def atualizar_excel_validacao(
    resultados_validacao: List[Dict[str, Any]],
    arquivo_excel: Path = None
) -> None:
    """
    Atualiza aba VALIDACAO_RETROATIVA no Excel com resultados.
    
    Args:
        resultados_validacao: Lista de dicionários com resultados
        arquivo_excel: Path para o Excel (default: ANALISE_HISTORICO_COMPLETO.xlsx)
    """
    if arquivo_excel is None:
        arquivo_excel = RESULTADO_DIR / 'ANALISE_HISTORICO_COMPLETO.xlsx'
    
    print(f"\n💾 Atualizando {arquivo_excel.name}...")
    
    # Preparar dados para DataFrame
    registros = []
    for res in resultados_validacao:
        registro = {
            'Concurso': res['concurso'],
            'Data': res['data'],
            'Numeros_Sorteados': "-".join(f"{n:02d}" for n in sorted(res['numeros_sorteados'])),
            
            # TOP 30
            'TOP_30': "-".join(f"{n:02d}" for n in res['validacao']['top_30']['numeros'][:30]),
            'Acertos_30': res['validacao']['top_30']['acertos'],
            'Taxa_30': f"{res['validacao']['top_30']['taxa_acerto']*100:.1f}%",
            
            # TOP 20
            'TOP_20': "-".join(f"{n:02d}" for n in res['validacao']['top_20']['numeros'][:20]),
            'Acertos_20': res['validacao']['top_20']['acertos'],
            
            # TOP 10
            'TOP_10': "-".join(f"{n:02d}" for n in res['validacao']['top_10']['numeros'][:10]),
            'Acertos_10': res['validacao']['top_10']['acertos'],
            
            # TOP 9
            'TOP_9': "-".join(f"{n:02d}" for n in res['validacao']['top_9']['numeros'][:9]),
            'Acertos_9': res['validacao']['top_9']['acertos'],
            
            # IA
            'Usou_IA': 'Sim' if res.get('ia_sugestoes') else 'Não',
            'Acertos_IA_30': res.get('reavaliacao', {}).get('acertos_top_30_ia', '-'),
            'Melhoria': res.get('reavaliacao', {}).get('acertos_top_30_ia', 0) - res['validacao']['top_30']['acertos'] if res.get('reavaliacao') else 0
        }
        registros.append(registro)
    
    df_novo = pd.DataFrame(registros)
    
    # Salvar no Excel
    try:
        from openpyxl import load_workbook
        
        if arquivo_excel.exists():
            wb = load_workbook(arquivo_excel)
        else:
            from openpyxl import Workbook
            wb = Workbook()
        
        # Remover aba se existir
        if 'VALIDACAO_RETROATIVA' in wb.sheetnames:
            del wb['VALIDACAO_RETROATIVA']
        
        # Criar nova aba
        ws = wb.create_sheet('VALIDACAO_RETROATIVA')
        
        # Escrever dados
        from openpyxl.utils.dataframe import dataframe_to_rows
        for row in dataframe_to_rows(df_novo, index=False, header=True):
            ws.append(row)
        
        wb.save(arquivo_excel)
        print(f"   ✅ Aba VALIDACAO_RETROATIVA atualizada com {len(registros)} registros")
        
    except Exception as e:
        print(f"   ⚠️  Erro ao atualizar Excel: {e}")


# ============================================================================
# EM DESENVOLVIMENTO: FASE 5 - ANÁLISE DE GRUPOS
# (Será implementada no próximo arquivo)
# ============================================================================



"""
[CONTINUAÇÃO] Validador Retroativo Expandido - MegaCLI v6.1
Função principal de execução e integração
"""

# (Continação do validador_retroativo_v2.py)

# Adicionar imports para a segunda parte
from src.validacao.validador_retroativo import (
    ler_ultimo_sorteio,
    extrair_previsoes_excel,
    extrair_ranking_indicadores
)
from src.validacao.analisador_grupos_indicadores import analisar_grupos_indicadores


# ============================================================================
# FUNÇÃO PRINCIPAL DE EXECUÇÃO
# ============================================================================

def executar_validacao_completa_v2(n_ultimos_sorteios: int = 5, usar_ia: bool = True, analisar_grupos: bool = False) -> Dict[str, Any]:
    """
    Função principal que orquestra toda a validação retroativa expandida.
    
    Args:
        n_ultimos_sorteios: Número de últimos sorteios a analisar
        usar_ia: Se True, consulta IA para análise de indicadores
        analisar_grupos: Se True, executa análise combinatória de grupos
        
    Returns:
        Dicionário com resultado completo da validação
    """
    print("\\n" + "="*80)
    print("🚀 VALIDAÇÃO RETROATIVA INTELIGENTE v2.0")
    print("="*80 + "\\n")
    
    try:
        # 1. Ler dados
        print("📂 Passo 1: Lendo dados...")
        df_sorteios, nome_arquivo = ler_ultimo_sorteio()
        df_ranking = extrair_ranking_indicadores()
        
        # Converter ranking para formato de lista de dicts
        ranking_list = []
        for _, row in df_ranking.iterrows():
            ranking_list.append({
                'indicador': row['Indicador'],
                'relevancia': row.get('Peso_Atual', row.get('Relevancia', 50.0))
            })
        
        print(f"   ✅ {len(df_sorteios)} sorteios carregados")
        print(f"   ✅ {len(ranking_list)} indicadores no ranking\\n")
        
        # 2. Processar últimos N sorteios
        print(f"🎯 Passo 2: Validando últimos {n_ultimos_sorteios} sorteios...\\n")
        
        resultados_validacao = []
        inicio = max(150, len(df_sorteios) - n_ultimos_sorteios)  # Margem para indicadores
        
        for i in range(inicio, len(df_sorteios)):
            concurso = df_sorteios.iloc[i]['Concurso']
            data = df_sorteios.iloc[i].get('Data Sorteio', df_sorteios.iloc[i].get('Data', 'N/A'))
            
            # Extrair números sorteados
            numeros_sorteados = [int(df_sorteios.iloc[i][f'Bola{k}']) for k in range(1, 7)]
            
            print(f"   📍 Concurso {concurso} ({data})")
            print(f"      Números: {'-'.join(f'{n:02d}' for n in sorted(numeros_sorteados))}")
            
            # Histórico até este sorteio
            df_historico_ate_sorteio = df_sorteios.iloc[:i]
            
            # 2.1 Validação multi-nível
            validacao = validar_multi_nivel(numeros_sorteados, df_historico_ate_sorteio, ranking_list)
            
            print(f"      ✅ TOP 30: {validacao['top_30']['acertos']}/6 acertos")
            print(f"      ✅ TOP 9: {validacao['top_9']['acertos']}/6 acertos")
            
            resultado_sorteio = {
                'concurso': concurso,
                'data': str(data),
                'numeros_sorteados': numeros_sorteados,
                'validacao': validacao
            }
            
            # 2.2 Consultar IA (se habilitado e não acertou 100%)
            if usar_ia and validacao['top_30']['acertos'] < 6:
                sugestoes = consultar_ia_analise_retroativa(
                    concurso,
                    numeros_sorteados,
                    validacao,
                    ranking_list
                )
                
                if sugestoes:
                    resultado_sorteio['ia_sugestoes'] = sugestoes
                    
                    # 2.3 Reavaliar com IA
                    reavaliacao = reavaliar_com_ia(
                        numeros_sorteados,
                        df_historico_ate_sorteio,
                        ranking_list,
                        sugestoes
                    )
                    
                    resultado_sorteio['reavaliacao'] = reavaliacao
                    
                    melhoria = reavaliacao['acertos_top_30_ia'] - validacao['top_30']['acertos']
                    if melhoria > 0:
                        print(f"      🌟 Com IA: +{melhoria} acertos! ({reavaliacao['acertos_top_30_ia']}/6)")
                    elif melhoria < 0:
                        print(f"      ⚠️  Com IA: {melhoria} acertos ({reavaliacao['acertos_top_30_ia']}/6)")
                    else:
                        print(f"      ➡️  Com IA: mesmo resultado ({reavaliacao['acertos_top_30_ia']}/6)")
            
            resultados_validacao.append(resultado_sorteio)
            print()
        
        # 3. Análise de grupos (se habilitado)
        melhores_grupos = None
        if analisar_grupos:
            print("\\n🔬 Passo 3: Analisando grupos de indicadores...")
            melhores_grupos = analisar_grupos_indicadores(
                df_sorteios,
                ranking_list,
                tamanho_grupo=5,
                n_jogos_teste=min(50, len(df_sorteios) // 2)
            )
            
            print("\\n📊 Top 5 Melhores Grupos de Indicadores:")
            for i, grupo in enumerate(melhores_grupos[:5], 1):
                print(f"   {i}. Taxa: {grupo['taxa_acerto_top30']*100:.1f}% - {', '.join(grupo['grupo'])}")
        
        # 4. Atualizar Excel
        print("\\n💾 Passo 4: Atualizando Excel...")
        atualizar_excel_validacao(resultados_validacao)
        
        # 5. Gerar relatório final
        print("\\n" + "="*80)
        print("📈 RESUMO DA VALIDAÇÃO")
        print("="*80)
        
        total_analisados = len(resultados_validacao)
        acertos_30 = [r['validacao']['top_30']['acertos'] for r in resultados_validacao]
        acertos_9 = [r['validacao']['top_9']['acertos'] for r in resultados_validacao]
        
        print(f"\\nSorteios analisados: {total_analisados}")
        print(f"\\nTOP 30:")
        print(f"   • Média de acertos: {np.mean(acertos_30):.2f}/6")
        print(f"   • Taxa 4+ acertos: {sum(a >= 4 for a in acertos_30)/total_analisados*100:.1f}%")
        print(f"\\nTOP 9:")
        print(f"   • Média de acertos: {np.mean(acertos_9):.2f}/6")
        print(f"   • Taxa 3+ acertos: {sum(a >= 3 for a in acertos_9)/total_analisados*100:.1f}%")
        
        if usar_ia:
            melhorias = [r.get('reavaliacao', {}).get('acertos_top_30_ia', r['validacao']['top_30']['acertos']) - r['validacao']['top_30']['acertos'] 
                         for r in resultados_validacao]
            print(f"\\n🤖 IA:")
            print(f"   • Melhoria média: {np.mean([m for m in melhorias if m != 0]):.2f} números")
            print(f"   • Jogos melhorados: {sum(m > 0 for m in melhorias)}/{total_analisados}")
        
        print("\\n" + "="*80)
        print("✅ Validação concluída com sucesso!")
        print("="*80 + "\\n")
        
        analise_completa = {
            'timestamp': datetime.now().isoformat(),
            'arquivo_sorteios': nome_arquivo,
            'n_sorteios_analisados': total_analisados,
            'resultados': resultados_validacao,
            'melhores_grupos': melhores_grupos,
            'resumo': {
                'media_acertos_top30': np.mean(acertos_30),
                'media_acertos_top9': np.mean(acertos_9),
                'taxa_sucesso_top30': sum(a >= 4 for a in acertos_30)/total_analisados,
                'melhorias_ia': melhorias if usar_ia else []
            }
        }
        
        # Salvar JSON
        arquivo_json = RESULTADO_DIR / 'validacao_retroativa' / f'analise_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'
        arquivo_json.parent.mkdir(exist_ok=True)
        
        with open(arquivo_json, 'w', encoding='utf-8') as f:
            json.dump(analise_completa, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"📁 Análise detalhada salva em: {arquivo_json.name}\\n")
        
        return analise_completa
        
    except Exception as e:
        print(f"\\n❌ Erro durante validação: {e}")
        import traceback
        traceback.print_exc()
        return None


# Exports
__all__ = [
    'validar_multi_nivel',
    'consultar_ia_analise_retroativa',
    'reavaliar_com_ia',
    'atualizar_excel_validacao',
    'executar_validacao_completa_v2'
]


# Teste standalone
if __name__ == "__main__":
    print("\\n🧪 Testando Validador Retroativo v2.0...\\n")
    resultado = executar_validacao_completa_v2(
        n_ultimos_sorteios=3,
        usar_ia=True,
        analisar_grupos=False
    )
    
    if resultado:
        print("\\n✅ Teste concluído com sucesso!")
    else:
        print("\\n❌ Erro no teste!")
