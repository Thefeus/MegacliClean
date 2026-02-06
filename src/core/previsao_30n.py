"""
Sistema de Previsão TOP_30N - MegaCLI v6.0

Sistema especializado para prever o próximo sorteio da Mega-Sena.
Analisa últimos 500 jogos, seleciona 30 números mais prováveis
e refina para os 10 melhores para apostar.

Autor: MegaCLI Team
Data: 23/01/2026
Versão: 1.0.0
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Tuple
from collections import Counter
from src.core.sistema_voto import coletar_votos_indicadores, gerar_analise_intersecao


def analisar_ultimos_500_jogos(df_historico: pd.DataFrame) -> Dict[str, any]:
    """
    Analisa últimos 500 sorteios do histórico.
    
    Args:
        df_historico: DataFrame com histórico completo
        
    Returns:
        Dicionário com estatísticas dos últimos 500 jogos
    """
    df_500 = df_historico.tail(500)
    
    # Extrair todos os números sorteados
    todos_numeros = []
    for _, row in df_500.iterrows():
        numeros = [row[f'Bola{i}'] for i in range(1, 7)]
        todos_numeros.extend(numeros)
    
    # Calcular frequências
    frequencias = Counter(todos_numeros)
    
    # Estatísticas
    stats = {
        'total_jogos': len(df_500),
        'primeiro_concurso': df_500.iloc[0]['Concurso'],
        'ultimo_concurso': df_500.iloc[-1]['Concurso'],
        'frequencias': frequencias,
        'numeros_mais_frequentes': frequencias.most_common(30)
    }
    
    return stats


def calcular_score_numero(
    numero: int,
    df_historico: pd.DataFrame,
    janela: int = 500
) -> float:
    """
    Calcula score de probabilidade para um número.
    
    Args:
        numero: Número a avaliar (1-60)
        df_historico: DataFrame com histórico
        janela: Janela de análise
        
    Returns:
        Score de 0-100
    """
    df_recente = df_historico.tail(janela)
    
    # 1. Frequência geral
    freq_total = 0
    for _, row in df_recente.iterrows():
        numeros = [row[f'Bola{i}'] for i in range(1, 7)]
        if numero in numeros:
            freq_total += 1
    
    score_freq = (freq_total / janela) * 100
    
    # 2. Frequência recente (últimos 50)
    df_50 = df_recente.tail(50)
    freq_recente = 0
    for _, row in df_50.iterrows():
        numeros = [row[f'Bola{i}'] for i in range(1, 7)]
        if numero in numeros:
            freq_recente += 1
    
    bonus_recencia = (freq_recente / 50) * 20
    
    # 3. Tendência (últimos 100 vs anteriores)
    df_100_recentes = df_recente.tail(100)
    df_100_anteriores = df_recente.iloc[-200:-100] if len(df_recente) >= 200 else df_recente.iloc[:-100]
    
    freq_100_rec = sum(1 for _, row in df_100_recentes.iterrows() 
                       if numero in [row[f'Bola{i}'] for i in range(1, 7)])
    freq_100_ant = sum(1 for _, row in df_100_anteriores.iterrows() 
                       if numero in [row[f'Bola{i}'] for i in range(1, 7)])
    
    tendencia = ((freq_100_rec - freq_100_ant) / 100) * 10
    
    # Score final
    score_final = score_freq * 0.40 + bonus_recencia + tendencia
    
    return max(0, min(100, score_final))


def selecionar_top_30_numeros(
    df_historico: pd.DataFrame,
    ranking_indicadores: List[Dict],
    verbose: bool = True
) -> Tuple[List[int], Dict[int, float], List[Dict], List[Dict]]:
    """
    Seleciona os 30 números mais prováveis usando Sistema de Votação Rastreável.
    
    Args:
        df_historico: DataFrame com histórico
        ranking_indicadores: Ranking de indicadores (Pesos IA)
        verbose: Se True, exibe informações
        
    Returns:
        Tupla (
            lista de 30 números, 
            dicionário {número: score},
            lista de rastreabilidade (votos por indicador),
            lista de interseção (detalhe por número)
        )
    """
    if verbose:
        print("\n🎯 Selecionando TOP 30 Números (Sistema de Voto Rastreável)")
        print("="*70)
    
    # 1. Coletar Votos (O Coração da Rastreabilidade)
    scores_counter, lista_rastreabilidade = coletar_votos_indicadores(df_historico, ranking_indicadores)
    
    # Converter para dict
    scores = dict(scores_counter)
    
    # Ordenar e pegar top 30
    numeros_ordenados = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    top_30 = [num for num, _ in numeros_ordenados[:30]]
    scores_top_30 = {num: score for num, score in numeros_ordenados[:30]}
    
    # 2. Gerar Análise de Interseção (Justificativa) para o Top 30
    analise_intersecao = gerar_analise_intersecao(top_30, lista_rastreabilidade)
    
    if verbose:
        print(f"\n📊 TOP 30 Números Selecionados:")
        print(f"\n{'#':<4} {'Número':<8} {'Score':<10} {'Barra':<30}")
        print("-"*70)
        
        max_score = max(scores_top_30.values()) if scores_top_30 else 1
        for i, (num, score) in enumerate(numeros_ordenados[:30], 1):
            barra_len = int((score / max_score) * 25)
            barra = '█' * barra_len
            print(f"{i:<4} {num:02d}       {score:>6.2f}     {barra}")
            
    return sorted(top_30), scores_top_30, lista_rastreabilidade, analise_intersecao


def refinar_selecao(
    numeros_base: List[int],
    scores_base: Dict[int, float],
    df_historico: pd.DataFrame,
    verbose: bool = True
) -> Tuple[List[int], Dict[int, float]]:
    """
    Refina a seleção de números usando co-ocorrência e diversidade.
    Retorna a lista completa ordenada por score refinado.
    
    Args:
        numeros_base: Lista de números para refinar (ex: top 30)
        scores_base: Scores originais dos números
        df_historico: DataFrame com histórico
        
    Returns:
        Tupla (lista ordenada completa, dicionário {número: score refinado})
    """
    if verbose:
        print(f"\n⭐ Refinando seleção de {len(numeros_base)} números...")
        print("="*70)
        
    if not numeros_base:
        return [], {}
    
    # Analisar co-ocorrências entre os números base
    df_500 = df_historico.tail(500)
    
    # Matriz de co-ocorrência
    matriz_co_oc = np.zeros((61, 61))
    for _, row in df_500.iterrows():
        numeros = [row[f'Bola{i}'] for i in range(1, 7)]
        for i, n1 in enumerate(numeros):
            for n2 in numeros[i+1:]:
                if n1 in numeros_base and n2 in numeros_base:
                    matriz_co_oc[n1][n2] += 1
                    matriz_co_oc[n2][n1] += 1
    
    # Calcular score refinado
    scores_refinados = {}
    for num in numeros_base:
        # Score base (se não tiver, assume 0 ou média)
        score_base = scores_base.get(num, 50.0)
        
        # Bônus por co-ocorrência com outros números do grupo
        co_oc_total = sum(matriz_co_oc[num][n] for n in numeros_base if n != num)
        # Normalizar co-ocorrência pelo tamanho do grupo
        bonus_co_oc = (co_oc_total / 500) * 10
        
        # Bônus por diversidade (dezena)
        bonus_diversidade = 2.0  # Todos ganham bônus base
        
        # Score final refinado
        score_refinado = score_base + bonus_co_oc + bonus_diversidade
        scores_refinados[num] = score_refinado
    
    # Ordenar por score refinado
    numeros_ordenados = sorted(scores_refinados.items(), key=lambda x: x[1], reverse=True)
    lista_refinada = [num for num, _ in numeros_ordenados]
    scores_finais = {num: score for num, score in numeros_ordenados}
    
    if verbose:
        print(f"\n🎲 TOP 10 REFINADOS (PREVIEW):")
        print(f"\n{'#':<4} {'Número':<8} {'Score':<10} {'Dezena':<8}")
        print("-"*70)
        
        for i, (num, score) in enumerate(numeros_ordenados[:10], 1):
            dezena = f"{(num//10)*10}-{(num//10)*10+9}"
            print(f"{i:<4} {num:02d}       {score:>6.2f}     {dezena}")
    
    return lista_refinada, scores_finais


def gerar_previsao_proximo_sorteio(
    df_historico: pd.DataFrame,
    ranking_indicadores: List[Dict] = None,
    verbose: bool = True
) -> Dict[str, any]:
    """
    Gera previsão completa para o próximo sorteio.
    
    Args:
        df_historico: DataFrame com histórico
        ranking_indicadores: Ranking de indicadores (opcional)
        verbose: Se True, exibe informações
        
    Returns:
        Dicionário com previsão completa
    """
    # Determinar próximo sorteio dinamicamente
    ultimo_sorteio = df_historico['Concurso'].max()
    proximo_sorteio = ultimo_sorteio + 1
    
    # Criar ranking simplificado se não fornecido
    if ranking_indicadores is None:
        ranking_indicadores = [
            {'indicador': f'Indicador{i}', 'relevancia': 100 - i*5}
            for i in range(1, 11)
        ]
    
    if verbose:
        print("\n" + "="*70)
        print(f"🔮 PREVISÃO PRÓXIMO SORTEIO - SISTEMA TOP_30N")
        print("="*70)
    
    # Análise base
    stats = analisar_ultimos_500_jogos(df_historico)
    
    if verbose:
        print(f"\n📊 Análise Base:")
        print(f"   • Jogos analisados: {stats['total_jogos']}")
        print(f"   • Período: Concurso {stats['primeiro_concurso']} a {stats['ultimo_concurso']}")
        print(f"   • Último sorteio: {ultimo_sorteio}")
        print(f"   • Próximo sorteio: {proximo_sorteio}")
    
    # Selecionar TOP 30 (retorna 4 valores)
    numeros_30, scores_30, _, _ = selecionar_top_30_numeros(
        df_historico,
        ranking_indicadores,
        verbose=verbose
    )
    
    # Refinar para TOP 10
    # Refinar para TOP 10 usando a funcao refinar_selecao disponível
    # Nota: refinar_selecao retorna a lista toda, pegaremos os top 10 depois
    lista_refinada, scores_refinados = refinar_selecao(
        numeros_30,
        scores_30,
        df_historico,
        verbose=verbose
    )
    
    numeros_10 = lista_refinada[:10]
    scores_10 = {n: scores_refinados[n] for n in numeros_10}
    
    # Calcular métricas
    score_medio = np.mean(list(scores_10.values()))
    
    if score_medio >= 85:
        confianca = "MUITO ALTA"
    elif score_medio >= 75:
        confianca = "ALTA"
    elif score_medio >= 65:
        confianca = "MÉDIA"
    else:
        confianca = "BAIXA"
    
    # Resultado
    resultado = {
        'concurso_previsto': proximo_sorteio,
        'top_30': numeros_30,
        'scores_30': scores_30,
        'top_10_aposta': numeros_10,
        'scores_10': scores_10,
        'score_medio': score_medio,
        'confianca': confianca,
        'stats': stats
    }
    
    if verbose:
        print(f"\n📈 Métricas da Previsão:")
        print(f"   • Score médio: {score_medio:.2f}")
        print(f"   • Confiança: {confianca}")
        print(f"   • Diversidade: {len(set(n//10 for n in numeros_10))} dezenas")
    
    # Exportar para planilha principal
    try:
        from pathlib import Path
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        # Caminho da planilha principal
        PROJECT_ROOT = Path(__file__).parent.parent.parent
        excel_path = PROJECT_ROOT / 'Resultado' / 'ANALISE_HISTORICO_COMPLETO.xlsx'
        
        if verbose:
            print(f"\n💾 Exportando para planilha principal...")
        
        # Carregar ou criar planilha
        if excel_path.exists():
            wb = load_workbook(str(excel_path))
        else:
            from openpyxl import Workbook
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        
        # Remover aba PREVISÕES se existir
        if 'PREVISÕES' in wb.sheetnames:
            del wb['PREVISÕES']
        
        # Criar nova aba
        ws = wb.create_sheet('PREVISÕES', 0)  # Primeira posição
        
        # Estilos
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Título
        ws['A1'] = f'PREVISÃO SORTEIO {proximo_sorteio} - SISTEMA TOP_30N'
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        ws.merge_cells('A1:G1')
        
        # Informações gerais
        row = 3
        ws[f'A{row}'] = 'Data da Previsão:'
        ws[f'B{row}'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        ws[f'B{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = 'Sorteio Previsto:'
        ws[f'B{row}'] = proximo_sorteio
        ws[f'B{row}'].font = Font(bold=True, color='0000FF')
        
        row += 1
        ws[f'A{row}'] = 'Confiança:'
        ws[f'B{row}'] = confianca
        ws[f'B{row}'].font = Font(bold=True, color='FF0000' if confianca == 'BAIXA' else '00B050')
        
        row += 1
        ws[f'A{row}'] = 'Score Médio:'
        ws[f'B{row}'] = f'{score_medio:.2f}'
        
        row += 1
        ws[f'A{row}'] = 'Jogos Analisados:'
        ws[f'B{row}'] = stats['total_jogos']
        
        # TOP 10 - Aposta Final
        row += 3
        ws[f'A{row}'] = '🎯 APOSTA FINAL (TOP 10)'
        ws[f'A{row}'].font = Font(bold=True, size=12, color='FF0000')
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        headers = ['Rank', 'Número', 'Score', 'Frequência', 'Co-ocorrência', 'Tendência', 'Dezena']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        
        # Dados TOP 10
        for i, (num, score) in enumerate(sorted(scores_10.items(), key=lambda x: x[1], reverse=True), 1):
            row += 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=num)
            ws.cell(row=row, column=3, value=round(score, 2))
            ws.cell(row=row, column=4, value='-')  # Placeholder
            ws.cell(row=row, column=5, value='-')
            ws.cell(row=row, column=6, value='-')
            ws.cell(row=row, column=7, value=f"{(num//10)*10}-{(num//10)*10+9}")
            
            # Centralizar
            for col in range(1, 8):
                ws.cell(row=row, column=col).alignment = center_align
        
        # Aposta formatada
        row += 2
        aposta_str = '-'.join(f'{n:02d}' for n in sorted(numeros_10))
        ws[f'A{row}'] = 'Aposta:'
        ws[f'B{row}'] = aposta_str
        ws[f'B{row}'].font = Font(bold=True, size=14, color='0000FF')
        ws.merge_cells(f'B{row}:G{row}')
        
        # TOP 30 - Números Selecionados
        row += 3
        ws[f'A{row}'] = '📊 TOP 30 NÚMEROS SELECIONADOS'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws.merge_cells(f'A{row}:G{row}')
        
        row += 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = center_align
        
        # Dados TOP 30
        for i, (num, score) in enumerate(sorted(scores_30.items(), key=lambda x: x[1], reverse=True), 1):
            row += 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=num)
            ws.cell(row=row, column=3, value=round(score, 2))
            ws.cell(row=row, column=4, value='-')
            ws.cell(row=row, column=5, value='-')
            ws.cell(row=row, column=6, value='-')
            ws.cell(row=row, column=7, value=f"{(num//10)*10}-{(num//10)*10+9}")
            
            for col in range(1, 8):
                ws.cell(row=row, column=col).alignment = center_align
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 16
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        
        # Salvar
        wb.save(str(excel_path))
        
        if verbose:
            print(f"✅ Previsão salva em: {excel_path}")
            print(f"   Aba: PREVISÕES")
        
    except PermissionError:
        if verbose:
            print(f"\n❌ ERRO DE PERMISSÃO: Não foi possível salvar o arquivo.")
            print(f"   Por favor, FECHE o arquivo Excel '{excel_path.name}' e tente novamente.")
    except Exception as e:
        if verbose:
            print(f"⚠️  Erro ao exportar para Excel: {e}")
    
    return resultado


# Exports
__all__ = [
    'analisar_ultimos_500_jogos',
    'selecionar_top_30_numeros',
    'refinar_selecao',
    'gerar_previsao_proximo_sorteio'
]


# Teste standalone
if __name__ == "__main__":
    print("\n🧪 Testando Sistema TOP_30N...\n")
    
    import sys
    from pathlib import Path
    PROJECT_ROOT = Path(__file__).parent.parent.parent
    sys.path.insert(0, str(PROJECT_ROOT))
    
    from src.core.config import ARQUIVO_HISTORICO
    
    # Carregar histórico
    df_historico = pd.read_excel(str(ARQUIVO_HISTORICO), sheet_name='MEGA SENA')
    print(f"✅ {len(df_historico)} sorteios carregados")
    
    # Ranking de teste
    ranking_teste = [
        {'indicador': f'Ind{i}', 'relevancia': 100-i*5}
        for i in range(1, 11)
    ]
    
    # Gerar previsão
    resultado = gerar_previsao_proximo_sorteio(df_historico, ranking_teste, verbose=True)
    
    print(f"\n✅ Previsão gerada com sucesso!")
    print(f"\n🎲 APOSTA FINAL PARA SORTEIO {resultado['concurso_previsto']}:")
    print(f"   {'-'.join(f'{n:02d}' for n in resultado['top_10_aposta'])}")
    print(f"   Confiança: {resultado['confianca']}\n")

