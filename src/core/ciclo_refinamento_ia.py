"""
Ciclo de Refinamento Iterativo com IA - MegaCLI v6.1

Módulo orquestrador que executa o loop de melhoria contínua:
1. Consulta a IA para ajustar pesos dos indicadores.
2. Atualiza o ranking de indicadores.
3. Gera previsão (Top 30) para 1000 jogos históricos (backtest).
4. Mede a performance (quantos números certos dentro do Top 30).
5. Relata a evolução.

Autor: MegaCLI Team
Data: 23/01/2026
"""


import pandas as pd
from typing import List, Dict, Any
from tqdm import tqdm
import time

# Imports do sistema
from src.utils.consultar_ia_refinamento import obter_sugestao_pesos
from src.core.previsao_30n import selecionar_top_30_numeros, refinar_selecao
from src.validacao.validador_1000_jogos import contar_acertos

from src.core.config import ARQUIVO_HISTORICO, RESULTADO_DIR
from src.core.analise_params import AnaliseConfig
from datetime import datetime
import os
import json
from pathlib import Path

ARQUIVO_ESTADO_PREVISAO = RESULTADO_DIR / 'estado_proxima_previsao.json'

def carregar_ranking_excel() -> pd.DataFrame:
    """Carrega o ranking atual do Excel."""
    arquivo_excel = RESULTADO_DIR / 'ANALISE_HISTORICO_COMPLETO.xlsx'

    df_ranking = pd.DataFrame()
    if os.path.exists(arquivo_excel):
        try:
            print(f"   📖 Lendo arquivo: {arquivo_excel}")
            # Tentar ler o Ranking atual
            try:
                df_ranking = pd.read_excel(arquivo_excel, sheet_name='RANKING INDICADORES')
                print(f"   📑 Aba RANKING: {len(df_ranking)} indicadores encontrados.")
            except:
                print("   ⚠️  Aba RANKING INDICADORES não encontrada.")

            # VERIFICAÇÃO DE INTEGRIDADE (Recuperação de Desastre)
            # Se tivermos poucos indicadores (ex: erro anterior salvou apenas 4), tentar restaurar
            if len(df_ranking) < 15: # 42 é o esperado
                print(f"   ⚠️  Detectada possível corrupção do ranking (apenas {len(df_ranking)} registros). Tentando restaurar...")
                
                # 1. Tentar ler de LISTA INDICADORES
                try:
                    df_lista = pd.read_excel(arquivo_excel, sheet_name='LISTA INDICADORES')
                    if not df_lista.empty:
                        print(f"   ♻️  Restaurando a partir de LISTA INDICADORES ({len(df_lista)} registros)...")
                        # Merge para manter o que sobrou e adicionar o que falta
                        col_nome = next((c for c in df_lista.columns if c.lower() in ['indicador', 'nome']), None)
                        if col_nome:
                            # Preparar dataframe base com nomes
                            df_base = pd.DataFrame({ 'Indicador': df_lista[col_nome] })
                            if 'Indicador' in df_ranking.columns:
                                df_ranking = pd.merge(df_base, df_ranking, on='Indicador', how='left')
                            else:
                                df_ranking = df_base
                            
                            # Preencher NaNs em Peso_Atual com valor padrão
                            if 'Peso_Atual' in df_ranking.columns:
                                df_ranking['Peso_Atual'] = df_ranking['Peso_Atual'].fillna(50.0)
                            else:
                                df_ranking['Peso_Atual'] = 50.0
                                
                            print(f"   ✅ Ranking restaurado para {len(df_ranking)} indicadores.")
                            return df_ranking
                except:
                    pass
                
                # 2. Se falhar, usar lista Hardcoded (Fallback Supremo)
                print("   ⚠️  LISTA INDICADORES não disponível. Usando lista padrão do sistema...")
                lista_padrao = [
                    'Quadrantes', 'Div9', 'Fibonacci', 'Div6', 'Mult5', 'Div3', 'Gap', 'Primos', 'Simetria', 'ParImpar', 'Amplitude', 'Soma',
                    'RaizDigital', 'VariacaoSoma', 'Conjugacao', 'RepeticaoAnterior', 'FrequenciaMensal',
                    'Sequencias', 'DistanciaMedia', 'NumerosExtremos', 'PadraoDezena', 'CicloAparicao',
                    'TendenciaQuadrantes', 'CiclosSemanais', 'AcumulacaoConsecutiva', 'JanelaDeslizante',
                    'MatrizPosicional', 'ClusterEspacial', 'SimetriaCentral',
                    'FrequenciaRelativa', 'DesvioFrequencia', 'EntrópiaDistribuicao', 'CorrelacaoTemporal',
                    'SomaDigitos', 'PadraoModular',
                    'ScoreAnomalia', 'ProbabilidadeCondicional', 'ImportanciaFeature',
                    'PadroesSubconjuntos', 'MicroTendencias', 'AnaliseContextual', 'Embedding'
                ]
                df_base = pd.DataFrame({'Indicador': lista_padrao})
                if not df_ranking.empty and 'Indicador' in df_ranking.columns:
                    df_ranking = pd.merge(df_base, df_ranking, on='Indicador', how='left')
                else:
                    df_ranking = df_base
                
                df_ranking['Peso_Atual'] = df_ranking.get('Peso_Atual', pd.Series([50.0]*len(df_ranking))).fillna(50.0)
                print(f"   ✅ Ranking restaurado via Hardcode ({len(df_ranking)} indicadores).")
                
            return df_ranking
            
        except Exception as e:
            print(f"   ❌ Erro ao ler Excel: {e}")
            pass
    return pd.DataFrame()

def gerar_analise_desvio(previsao_json: Dict, numeros_reais: set) -> Dict:
    """
    Compara a previsão anterior com o resultado real e gera insights.
    """
    sets_prev = {
        'Top 30': set(previsao_json.get('top_30', [])),
        'Top 20': set(previsao_json.get('top_20', [])),
        'Top 10': set(previsao_json.get('top_10', [])),
        'Top 9':  set(previsao_json.get('top_9', []))
    }
    
    analise = {
        'acertos': {},
        'desvios': [],
        'resumo_ia': ''
    }
    
    # 1. Contar acertos
    for label, nums_prev in sets_prev.items():
        acertos = numeros_reais & nums_prev
        analise['acertos'][label] = {
            'qtd': len(acertos),
            'numeros': sorted(list(acertos))
        }
        
    # 2. Analisar Desvios (Números que saíram mas não estavam na previsão)
    # Focando no Top 30 como base de corte principal
    nao_previstos = numeros_reais - sets_prev['Top 30']
    
    # Rastrear por que foram excluídos (se tivermos o detalhe)
    detalhes_numeros = previsao_json.get('detalhe_numeros', [])
    mapa_detalhes = {d['Numero']: d for d in detalhes_numeros} if detalhes_numeros else {}
    
    msg_desvio = []
    if nao_previstos:
        msg_desvio.append(f"⚠️  {len(nao_previstos)} números sorteados fora do TOP 30: {sorted(list(nao_previstos))}")
        for n in nao_previstos:
            if n in mapa_detalhes:
                info = mapa_detalhes[n]
                msg_desvio.append(f"   • Bola {n}: Score {info.get('Score_Total', 0):.1f}. Fontes: {info.get('Fontes_Voto', 'N/A')}")
            else:
                msg_desvio.append(f"   • Bola {n}: Não monitorado ou score muito baixo.")
    else:
        msg_desvio.append("✅ Todos os números sorteados estavam no TOP 30!")
        
    analise['texto_desvio'] = "\n".join(msg_desvio)
    
    return analise

def atualizar_excel_refinamento(
    ranking_atual: pd.DataFrame, 
    novos_pesos: Dict[str, float],
    performance_ciclo: float = 0.0,
    dados_analise: Dict = None,
    dados_predicao: Dict = None,
    detalhes_validacao: List[Dict] = None,
    df_mega_hist: pd.DataFrame = None
) -> None:
    """
    Atualiza o Excel com os novos pesos, histórico, previsões e detalhes de validação.
    """
    arquivo_excel = RESULTADO_DIR / 'ANALISE_HISTORICO_COMPLETO.xlsx'
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    print(f"\n💾 Atualizando planilha: {arquivo_excel.name}...")
    
    # 1. Atualizar RANKING INDICADORES (Pesos)
    df_novo_ranking = ranking_atual.copy()
    
    # Detectar nomes das colunas (para compatibilidade com arquivos existentes)
    col_ind = 'Indicador'
    col_peso = 'Peso_Atual'
    
    if not df_novo_ranking.empty:
        # Tentar identificar coluna de nome
        for c in ['Indicador', 'indicador', 'Nome', 'nome']:
            if c in df_novo_ranking.columns:
                col_ind = c
                break
        
        # Tentar identificar coluna de peso/relevância
        for c in ['Peso_Atual', 'peso_atual', 'Relevancia', 'relevancia', 'Peso', 'peso']:
            if c in df_novo_ranking.columns:
                col_peso = c
                break
    
    # 2.A Garantir que a coluna de peso seja float para evitar FutureWarning
    if col_peso in df_novo_ranking.columns:
        df_novo_ranking[col_peso] = df_novo_ranking[col_peso].astype(float)
    
    if df_novo_ranking.empty:
        # Se vazio, criar estrutura básica
        df_novo_ranking = pd.DataFrame({
            col_ind: list(novos_pesos.keys()), 
            col_peso: list(novos_pesos.values())
        })
    else:
        # Atualizar pesos existentes e adicionar novos
        for ind, peso in novos_pesos.items():
            if ind in df_novo_ranking[col_ind].values:
                df_novo_ranking.loc[df_novo_ranking[col_ind] == ind, col_peso] = peso
            else:
                # Adicionar novo
                novo_row = {col_ind: ind, col_peso: peso}
                df_novo_ranking = pd.concat([df_novo_ranking, pd.DataFrame([novo_row])], ignore_index=True)
    
    # Ordenar por peso
    if col_peso in df_novo_ranking.columns:
        df_novo_ranking = df_novo_ranking.sort_values(by=col_peso, ascending=False)

    # 2.A.1 Enriquecer Ranking com Metadados (Descrição e Categoria)
    # Recuperando visual rico solicitado pelo usuário
    from src.utils.descricoes_indicadores import criar_dicionario_completo
    dict_meta = criar_dicionario_completo()
    dict_meta_lower = {k.lower(): v for k, v in dict_meta.items()}

    # Garantir que colunas existam
    if 'Descrição' not in df_novo_ranking.columns:
        df_novo_ranking['Descrição'] = '-'
    if 'Categoria' not in df_novo_ranking.columns:
        df_novo_ranking['Categoria'] = '-'

    for idx, row in df_novo_ranking.iterrows():
        ind_nome = str(row[col_ind])
        meta = dict_meta.get(ind_nome)
        if not meta:
            meta = dict_meta_lower.get(ind_nome.lower(), {})
            
        if meta and isinstance(meta, dict):
            df_novo_ranking.at[idx, 'Descrição'] = meta.get('descricao', '-')
            df_novo_ranking.at[idx, 'Categoria'] = meta.get('categoria', '-')
            
    # 2.B Preparar DF LISTA INDICADORES (REMOVIDO - CÓDIGO APARTADO)
    # A atualização desta aba agora é feita por src/ferramentas/atualizar_lista_indicadores.py
    # para evitar sobrescrita acidental ou conflitos.
    pass
    
    # 3. Atualizar HISTÓRICO_PESOS (Append)
    # Criar linha com timestamp e indicativo de performance
    dados_hist = {
        'Data': timestamp,
        'Performance_Ciclo': round(performance_ciclo, 4)
    }
    dados_hist.update(novos_pesos)
    df_hist_novo = pd.DataFrame([dados_hist])
    
    # Carregar histórico existente se houver
    df_historico_pesos = pd.DataFrame()
    if os.path.exists(arquivo_excel):
        try:
            df_historico_pesos = pd.read_excel(arquivo_excel, sheet_name='HISTÓRICO_PESOS')
        except:
            pass
            
    df_historico_pesos = pd.concat([df_historico_pesos, df_hist_novo], ignore_index=True)
    
    # 4. Preparar Dados ANÁLISE IA
    # Converter para DF para facilitar escrita
    registros_analise = []
    if dados_analise:
        analise_geral = dados_analise.get('analise_ciclo', '')
        justificativas = dados_analise.get('justificativas_top', [])
        
        # Registro Principal
        registros_analise.append({
            'Data': timestamp,
            'Tipo': 'GERAL',
            'Indicador': '-',
            'Conteudo': analise_geral,
            'Performance_Ciclo': round(performance_ciclo, 4)
        })
        
        # Registro de Justificativas
        for just in justificativas:
            registros_analise.append({
                'Data': timestamp,
                'Tipo': 'JUSTIFICATIVA',
                'Indicador': just.get('indicador', '-'),
                'Conteudo': just.get('motivo', '-'),
                'Performance_Ciclo': '-'
            })
            
    df_analise_ia_novo = pd.DataFrame(registros_analise)
    df_analise_ia_historico = pd.DataFrame()
    
    if os.path.exists(arquivo_excel):
        try:
            df_analise_ia_historico = pd.read_excel(arquivo_excel, sheet_name='ANÁLISE IA')
        except:
            pass
            
    df_analise_ia_final = pd.concat([df_analise_ia_historico, df_analise_ia_novo], ignore_index=True)

    # 5. Preparar Dados PREVISÕES (Substitui COMB IA)
    df_previsoes_final = pd.DataFrame()
    if dados_predicao:
        # Formatar sequências
        def fmt(lista): return '-'.join(f'{n:02d}' for n in sorted(lista))
        
        t30 = fmt(dados_predicao.get('top_30', []))
        t20 = fmt(dados_predicao.get('top_20', []))
        t10 = fmt(dados_predicao.get('top_10', []))
        t09 = fmt(dados_predicao.get('top_9', []))
        
        # Métricas
        mets = dados_predicao.get('metricas_validacao', {})
        
        # Top Indicadores (Formatado)
        top_inds = sorted(novos_pesos.items(), key=lambda x: x[1], reverse=True)[:10]
        top_inds_str = ', '.join([f"{k} ({v:.1f})" for k, v in top_inds])
        
        reg_prev = [{
            'Data': timestamp,
            'Concurso_Base': dados_predicao.get('concurso_base', 'N/A'),
            'Perf_Global': round(performance_ciclo, 4),
            'Top_Indicadores': top_inds_str, # Inserindo antes do Top_30
            'Top_30': t30,
            'T30_Quadras': mets.get(30, {}).get('4', 0),
            'T30_Quinas': mets.get(30, {}).get('5', 0),
            'T30_Senas': mets.get(30, {}).get('6', 0),
            'Top_20': t20,
            'T20_Quadras': mets.get(20, {}).get('4', 0),
            'T20_Quinas': mets.get(20, {}).get('5', 0),
            'T20_Senas': mets.get(20, {}).get('6', 0),
            'Top_10': t10,
            'T10_Quadras': mets.get(10, {}).get('4', 0),
            'T10_Senas': mets.get(10, {}).get('6', 0),
            'Top_09': t09,
            'T09_Quadras': mets.get(9, {}).get('4', 0),
            'T09_Senas': mets.get(9, {}).get('6', 0),
            'Analise_IA': dados_analise.get('analise_ciclo', '')[:150] if dados_analise else '-'
        }]
        
        df_prev_novo = pd.DataFrame(reg_prev)
        
        # Carregar histórico
        df_prev_antigo = pd.DataFrame()
        if os.path.exists(arquivo_excel):
            try:
                # Tenta ler PREVISÕES, se não existir tenta COMB IA para migrar? Melhor criar nova.
                df_prev_antigo = pd.read_excel(arquivo_excel, sheet_name='PREVISÕES')
            except:
                pass
        
        df_previsoes_final = pd.concat([df_prev_antigo, df_prev_novo], ignore_index=True)

    # 6. Preparar Dados DETALHE 1000 JOGOS
    df_detalhe_validacao = pd.DataFrame()
    if detalhes_validacao:
        df_detalhe_validacao = pd.DataFrame(detalhes_validacao)

    # 7. Preparar Dados RASTREAMENTO INDICADORES (Auditoria)
    df_rastreamento = pd.DataFrame()
    if dados_predicao and 'detalhe_numeros' in dados_predicao:
        # Converter a lista de detalhes (Por Número)
        df_rastreamento = pd.DataFrame(dados_predicao['detalhe_numeros'])
        # Reordenar colunas
        cols_order = ['Numero', 'Score_Total', 'Fontes_Voto']
        col_existentes = [c for c in cols_order if c in df_rastreamento.columns]
        df_rastreamento = df_rastreamento[col_existentes]
        
        # Adicionar seção de Votos por Indicador (Abaixo, ou em outra tabela? Vamos fazer concat vertical com separador)
        # Hack de display Excel: Adicionar linhas vazias e depois a outra tabela
        sep_row = pd.DataFrame([{'Numero': '---', 'Score_Total': '---', 'Fontes_Voto': 'DETALHE POR INDICADOR ---'}])
        
        df_votos = pd.DataFrame(dados_predicao.get('rastro_votos', []))
        # Formatar lista de números sugeridos para string
        if 'Numeros_Sugeridos' in df_votos.columns:
            df_votos['Numeros_Sugeridos'] = df_votos['Numeros_Sugeridos'].apply(lambda x: str(x))
            
        # Renomear para alinhar (visual apenas)
        df_votos = df_votos.rename(columns={'Indicador': 'Numero', 'Peso_IA': 'Score_Total', 'Numeros_Sugeridos': 'Fontes_Voto'})
        
        df_rastreamento = pd.concat([df_rastreamento, sep_row, df_votos], ignore_index=True)

    # 8. Preparar Dados ANÁLISE PRÓXIMO SORTEIO
    # Esta aba serve como um Dashboard para o operador

    # 8. Preparar Dados ANÁLISE PRÓXIMO SORTEIO
    # Esta aba serve como um Dashboard para o operador
    rows_proximo = []
    
    # Cabeçalho Status
    concurso_previsto = '?'
    if dados_predicao:
        try:
            conc_base = int(dados_predicao.get('concurso_base', 0))
            concurso_previsto = str(conc_base + 1)
        except:
            pass
            
    rows_proximo.append(['STATUS SISTEMA:', f'Previsão Gerada em {timestamp}'])
    rows_proximo.append(['PRÓXIMO CONCURSO:', concurso_previsto])
    rows_proximo.append(['', '']) # Espaço
    
    if dados_predicao:
        # Seção de Números Sugeridos
        rows_proximo.append(['🎯 NÚMEROS PREVISTOS (TOP 9)', fmt(dados_predicao.get('top_9', []))])
        rows_proximo.append(['🔹 TOP 10', fmt(dados_predicao.get('top_10', []))])
        rows_proximo.append(['🔹 TOP 20', fmt(dados_predicao.get('top_20', []))])
        rows_proximo.append(['🔹 TOP 30', fmt(dados_predicao.get('top_30', []))])
        rows_proximo.append(['', ''])
    
    # Seção de Validação (Espaço para o usuário ou preenchido pelo sistema)
    rows_proximo.append(['📝 VALIDAÇÃO RESULTADO REAL', ''])
    rows_proximo.append(['Concurso Real:', 'Preencher ou Aguardar Sistema'])
    rows_proximo.append(['Dezenas Sorteadas:', 'Preencher (ex: 01-02-03...)'])
    rows_proximo.append(['', ''])
    rows_proximo.append(['🔍 ANÁLISE DE DESVIO (Feedback Loop)', ''])
    
    # Se houver uma análise de desvio PREENCHIDA (passada como argumento extra ou calculada antes)
    # Por padrão, deixamos instruções
    if 'analise_desvio' in dados_predicao: # Hack: Injetar análise aqui se existir
        desvio = dados_predicao['analise_desvio']
        rows_proximo.append(['Status Validação:', '✅ CONCLUÍDA'])
        rows_proximo.append(['Acertos Top 30:', str(desvio.get('acertos', {}).get('Top 30', {}).get('qtd', 0))])
        rows_proximo.append(['Acertos Top 9:', str(desvio.get('acertos', {}).get('Top 9', {}).get('qtd', 0))])
        rows_proximo.append(['Detalhe Desvio:', desvio.get('texto_desvio', 'Sem desvios')])
    else:
        rows_proximo.append(['Status Validação:', 'AGUARDANDO SORTEIO'])
        rows_proximo.append(['Obs:', 'O sistema preencherá aqui automaticamente no próximo ciclo se o sorteio já tiver ocorrido.'])

    df_proximo_sorteio = pd.DataFrame(rows_proximo, columns=['Parâmetro', 'Valor'])

    # Salvar usando OPENPYXL direto para preservar outras abas e formatações
    print(f"   💾 Salvando em: {arquivo_excel.absolute()}")
    print(f"   📑 Atualizando abas: RANKING, HISTÓRICO, ANÁLISE IA, PREVISÕES")
    
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Carregar workbook existente ou criar novo
        if os.path.exists(arquivo_excel):
            wb = load_workbook(arquivo_excel)
        else:
            from openpyxl import Workbook
            wb = Workbook()
            # Remover sheet padrão
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        
        # Helper para atualizar aba
        def update_sheet(wb, sheet_name, df):
            if sheet_name in wb.sheetnames:
                # Limpar aba existente
                ws = wb[sheet_name]
                ws.delete_rows(1, ws.max_row)
            else:
                # Criar nova
                ws = wb.create_sheet(sheet_name)
                
            if not df.empty:
                # Usar utilitário oficial: Mais rápido e robusto
                for row in dataframe_to_rows(df, index=False, header=True):
                    ws.append(row)
                print(f"      - Aba '{sheet_name}' atualizada com {len(df)} registros.")
            else:
                print(f"      - Aba '{sheet_name}' ignorada (DataFrame vazio).")
                    
        # Atualizar as abas
        update_sheet(wb, 'RANKING INDICADORES', df_novo_ranking)
        # update_sheet(wb, 'LISTA INDICADORES', df_lista)  <-- REMOVIDO
        update_sheet(wb, 'HISTÓRICO_PESOS', df_historico_pesos)
        update_sheet(wb, 'ANÁLISE IA', df_analise_ia_final)
        
        if not df_previsoes_final.empty:
            update_sheet(wb, 'PREVISÕES', df_previsoes_final)
            
        if not df_detalhe_validacao.empty:
            update_sheet(wb, 'DETALHE 1000 JOGOS', df_detalhe_validacao) 
            
        if not df_rastreamento.empty:
            update_sheet(wb, 'RASTREAMENTO INDICADORES', df_rastreamento) # Nova aba de auditoria!
            
        if not df_proximo_sorteio.empty:
            update_sheet(wb, 'ANÁLISE PRÓXIMO SORTEIO', df_proximo_sorteio)
            # Ajustar larguras
            ws_prox = wb['ANÁLISE PRÓXIMO SORTEIO']
            ws_prox.column_dimensions['A'].width = 30
            ws_prox.column_dimensions['B'].width = 80
            
        if df_mega_hist is not None and not df_mega_hist.empty:
            update_sheet(wb, 'MEGA SENA HIST', df_mega_hist)
            print("   ✅ Aba 'MEGA SENA HIST' criada com sucesso!")
        
        wb.save(arquivo_excel)
        print("✅ Planilha atualizada com sucesso (Modo Seguro OpenPyXL)!")
        
    except PermissionError:
        print(f"\n❌ ERRO DE PERMISSÃO: O arquivo '{arquivo_excel.name}' parece estar aberto.")
        print("   ⚠️  FECHE O ARQUIVO NO EXCEL E TENTE NOVAMENTE.")
        
    except Exception as e:
        print(f"❌ Erro ao salvar Excel (OpenPyXL): {e}")
        import traceback
        traceback.print_exc()


def executar_ciclo_refinamento(df_historico: pd.DataFrame, n_ciclos: int = 1) -> None:
    """
    Executa o ciclo de refinamento iterativo.
    """
    print("\n" + "="*80)
    print("🤖 INICIANDO CICLO DE REFINAMENTO ITERATIVO COM IA")
    print("="*80)
    
    # 1. Carregar Estado Atual do Excel
    print("\n📂 Passo 1: Carregando dados atuais do Excel...")
    df_ranking_atual = carregar_ranking_excel()
    
    if not df_ranking_atual.empty:
        print(f"   ✅ Ranking carregado: {len(df_ranking_atual)} indicadores conhecidos.")
    else:
        print("   ⚠️  Ranking vazio ou não encontrado. IA começará do zero.")
    
    # 1.B Verificar Feedback de Previsão Passada (Feedback Loop)
    analise_desvio_passada = None
    if os.path.exists(ARQUIVO_ESTADO_PREVISAO):
        try:
            with open(ARQUIVO_ESTADO_PREVISAO, 'r') as f:
                estado_anterior = json.load(f)
            
            # Verificar se o concurso previsto já ocorreu
            conc_previsto = estado_anterior.get('concurso_alvo')
            # Buscar no histórico
            if conc_previsto and 'Concurso' in df_historico.columns:
                # Converte para int para garantir
                conc_previsto = int(conc_previsto)
                row_real = df_historico[df_historico['Concurso'] == conc_previsto]
                
                if not row_real.empty:
                    print(f"\n🔁 FEEDBACK LOOP: Validando previsão anterior (Conc. {conc_previsto})...")
                    # Extrair números reais
                    jogo_real = row_real.iloc[0]
                    numeros_reais = set([jogo_real[f'Bola{k}'] for k in range(1, 7)])
                    
                    # Gerar Análise
                    analise_desvio_passada = gerar_analise_desvio(estado_anterior, numeros_reais)
                    print(f"   📊 Resultado: {analise_desvio_passada.get('texto_desvio', '').splitlines()[0]}")
                    
                    # Opcional: Aqui poderíamos enviar esse desvio para a IA ajustar os pesos
                    # Por enquanto, apenas registramos para o humano ver
        except Exception as e:
            print(f"⚠️  Erro ao validar estado anterior: {e}")

    # 2. Consultar IA
    print("\n🧠 Passo 2: Analisando performance e definindo novos pesos...")
    
    # Simular dados de validação recente para contexto (poderia vir de arquivo também)
    df_val_dummy = pd.DataFrame({'Acertos': [3]*100}) 
    
    # Preparar texto de feedback (se houver)
    feedback_txt = analise_desvio_passada.get('texto_desvio', '') if analise_desvio_passada else None
    
    # AGORA RECEBE TUPLA (Pesos, DadosAnalise)
    novos_pesos, dados_analise_ia = obter_sugestao_pesos(
        df_ranking_atual, 
        df_val_dummy,
        feedback_texto=feedback_txt
    )
    
    if not novos_pesos:
        print("❌ Falha ao obter pesos da IA. Abortando ciclo.")
        return
        
    # INFO ADICIONAL SOLICITADA
    qtd_jogos = len(df_historico)
    prox_sorteio = "?"
    if 'Concurso' in df_historico.columns:
        try:
            prox_sorteio = int(df_historico.iloc[-1]['Concurso']) + 1
        except:
            pass
            
    print(f"   ✅ Quantidade de Jogos da Base Histórica: {qtd_jogos}")
    if 'Concurso' in df_historico.columns:
        print(f"   🔍 Faixa de Concursos Carregada: {df_historico['Concurso'].min()} a {df_historico['Concurso'].max()}")
    print(f"   ✅ Próximo Sorteio: {prox_sorteio}")

    # EXIBIR ANÁLISE DA IA
    if 'analise_ciclo' in dados_analise_ia:
        print(f"\n💡 ANÁLISE DA IA:\n{dados_analise_ia['analise_ciclo']}")
    
    if 'justificativas_top' in dados_analise_ia:
        print("\n🔍 Justificativas de Destaque:")
        for just in dados_analise_ia['justificativas_top'][:3]:
            print(f"   • {just.get('indicador')}: {just.get('motivo')}")

    print("\n📊 Pesos Sugeridos pela IA (Todos):")
    sorted_pesos = sorted(novos_pesos.items(), key=lambda x: x[1], reverse=True)
    
    for i, (k, v) in enumerate(sorted_pesos, 1):
        print(f"   {i:02d}. {k}: {v:.1f}")

    # 4. Criar Ranking Refinado para Validação
    ranking_refinado = [
        {'indicador': k, 'relevancia': v} for k, v in novos_pesos.items()
    ]
    
    # 5. Validação Histórica (Backtest) Expandida
    print("\n⏳ Passo 3: Executando Validação Histórica (BASE COMPLETA) em Múltiplos Níveis...")
    
    # amostra_validacao = 1000  <-- REMOVIDO! Agora é full.
    # inicio_val = max(0, len(df_historico) - amostra_validacao)
    # Para validar, precisamos de um minimo de histórico para calcular indicadores. 
    # Vamos assumir salto inicial definido na configuração.
    inicio_val = AnaliseConfig.VALIDACAO_OFFSET_INICIAL # Ignora os primeiros jogos para ter base mínima
    
    # Estrutura para métricas
    
    # Estrutura para métricas
    metricas = {
        30: {'6': 0, '5': 0, '4': 0, 'total': 0},
        20: {'6': 0, '5': 0, '4': 0, 'total': 0},
        10: {'6': 0, '5': 0, '4': 0, 'total': 0},
        9:  {'6': 0, '5': 0, '4': 0, 'total': 0}
    }
    
    detalhes_1000_jogos = []
    
    for i in tqdm(range(inicio_val, len(df_historico))):
        df_corte = df_historico.iloc[:i]
        jogo_real = df_historico.iloc[i]
        numeros_reais = set([jogo_real[f'Bola{k}'] for k in range(1, 7)])
        
        # Gerar previsão TOP 30
        # Agora retorna 4 valores!
        top_30, scores_30, _, _ = selecionar_top_30_numeros(
            df_corte,
            ranking_refinado,
            verbose=False
        )
        
        # Refinar Filtros (Subsets)
        lista_refinada, _ = refinar_selecao(top_30, scores_30, df_corte, verbose=False)
        
        # Definir subconjuntos
        sets_numeros = {
            30: set(lista_refinada[:30]),
            20: set(lista_refinada[:20]),
            10: set(lista_refinada[:10]),
            9:  set(lista_refinada[:9])
        }
        
        # Validar e Registrar Métricas
        # Tentar obter data de forma segura
        data_sorteio = jogo_real.get('Data Sorteio', jogo_real.get('Data', 'N/A'))
        
        # Copiar TODAS as colunas do jogo real para o registro (MEGA SENA HIST)
        registro_jogo = jogo_real.to_dict()
        
        # Garantir formato de data string para Excel
        registro_jogo['Data Sorteio'] = data_sorteio
        registro_jogo['Dezenas_Reais'] = '-'.join(f'{n:02d}' for n in sorted(numeros_reais))
        
        # Remover colunas duplicadas ou desnecessarias se houver
        # (Opcional)
        
        for qtd, numeros_set in sets_numeros.items():
            acertos = len(numeros_reais & numeros_set)
            metricas[qtd]['total'] += acertos
            
            # Atualizar estatísticas gerais
            if acertos == 6: metricas[qtd]['6'] += 1
            elif acertos == 5: metricas[qtd]['5'] += 1
            elif acertos == 4: metricas[qtd]['4'] += 1
            
            # Registrar detalhe do jogo
            registro_jogo[f'Top_{qtd}'] = '-'.join(f'{n:02d}' for n in sorted(numeros_set))
            registro_jogo[f'Acertos_{qtd}'] = acertos
            
        detalhes_1000_jogos.append(registro_jogo)

    # Calcular média baseada no total processado
    total_proc = len(detalhes_1000_jogos)
    media_30 = metricas[30]['total'] / total_proc if total_proc > 0 else 0
    
    print("\n" + "="*80)
    print(f"📈 RELATÓRIO DE PERFORMANCE DO CICLO (Amostra {total_proc} Jogos)")
    print("="*80)
    # ... (rest of print loop is fine)

    # ... (Prediction block logic is fine)

    # Criar DataFrame MEGA SENA HIST
    df_mega_hist_completo = pd.DataFrame(detalhes_1000_jogos)

    # 3. Atualizar Excel (Persistência) - MOVIDO PARA O FINAL
    # atualizar_excel_refinamento(...)
    for qtd in [30, 20, 10, 9]:
        m = metricas[qtd]
        print(f"   🔹 TOP {qtd}: Quadras: {m['4']} | Quinas: {m['5']} | Senas: {m['6']}")
    
    # 6. Gerar Previsão para o Próximo Concurso (PREVISÕES / COMB IA)
    print("\n🔮 Gerando Combinação Otimizada para o próximo sorteio...")
    try:
        # Previsão Final usando todo histórico
        # Captura os dados de rastreabilidade (rastro_votos, detalhe_numeros)
        top_30_final, scores_30_final, rastro_votos, detalhe_numeros = selecionar_top_30_numeros(df_historico, ranking_refinado, verbose=True)
        
        lista_final_refinada, _ = refinar_selecao(top_30_final, scores_30_final, df_historico, verbose=True)
        
        dados_predicao = {
            'concurso_base': df_historico.iloc[-1]['Concurso'] if 'Concurso' in df_historico.columns else 'N/A',
            'top_30': lista_final_refinada[:30],
            'top_20': lista_final_refinada[:20],
            'top_10': lista_final_refinada[:10],
            'top_9':  lista_final_refinada[:9],
            'metricas_validacao': metricas,
            'rastro_votos': rastro_votos,           # NOVO: Quem votou em quem
            'detalhe_numeros': detalhe_numeros      # NOVO: Justificativa por número
        }
        print(f"   ✅ Previsão gerada para Top 30, 20, 10 e 9.")
        print(f"   🎯 Aposta Sugerida (Top 9): {'-'.join(f'{n:02d}' for n in sorted(dados_predicao['top_9']))}")
        
        # INJETAR FEEDBACK PASSADO NA PLANILHA
        if analise_desvio_passada:
            dados_predicao['analise_desvio'] = analise_desvio_passada
            
        # SALVAR ESTADO PARA PRÓXIMO CICLO
        try:
            # Helper para converter tipos numpy/pandas
            def to_std_type(obj):
                if hasattr(obj, 'item'): return obj.item() # numpy types
                if hasattr(obj, 'tolist'): return obj.tolist() # numpy arrays
                return obj

            estado_salvar = {
                'concurso_base': str(dados_predicao['concurso_base']), # Forçar string
                'concurso_alvo': str(int(dados_predicao['concurso_base']) + 1) if str(dados_predicao['concurso_base']).isdigit() else None,
                'top_30': [int(n) for n in dados_predicao['top_30']],
                'top_20': [int(n) for n in dados_predicao['top_20']],
                'top_10': [int(n) for n in dados_predicao['top_10']],
                'top_9':  [int(n) for n in dados_predicao['top_9']],
                'detalhe_numeros': dados_predicao.get('detalhe_numeros', []),
                'timestamp_geracao': datetime.now().isoformat()
            }
            with open(ARQUIVO_ESTADO_PREVISAO, 'w') as f:
                json.dump(estado_salvar, f, indent=4)
            print(f"   💾 Estado da previsão salvo para validação futura em: {ARQUIVO_ESTADO_PREVISAO.name}")
        except Exception as e:
            print(f"   ⚠️  Erro ao salvar estado da previsão: {e}")
        
    except Exception as e:
        print(f"❌ Erro ao gerar previsão final: {e}")
        dados_predicao = {}

    # 3. Atualizar Excel (Persistência)
    atualizar_excel_refinamento(
        df_ranking_atual, 
        novos_pesos, 
        media_30, 
        dados_analise_ia, 
        dados_predicao, 
        detalhes_1000_jogos, # Esse aqui ainda é usado para 'DETALHE 1000 JOGOS' (nome legado, agora é FULL)
        df_mega_hist=df_mega_hist_completo
    )
    
    if media_30 > 4.0:
        print("\n✅ RESULTADO: EXCELENTE. IA convergiu bem.")
    elif media_30 > 3.0:
        print("\n⚠️  RESULTADO: BOM. Pode melhorar.")
    else:
        print("\n❌ RESULTADO: FRACO. Sugere-se novo ciclo.")

    print("\n" + "="*80)
