"""
Ferramenta de Restauração de Abas do Excel - Validação
MegaCLI v6.3.0

Este script verifica e recria as abas necessárias para o funcionamento
do Ciclo de Refinamento IA e Validação Retroativa, caso estejam faltando.

Abas gerenciadas:
- RANKING INDICADORES (Pesos iniciais)
- HISTÓRICO_PESOS (Registro vazio com cabeçalho)
- ANÁLISE IA (Registro vazio com cabeçalho)
"""

import pandas as pd
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

# Adicionar raiz ao path
PROJECT_ROOT = Path(__file__).parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.core.config import ARQUIVO_HISTORICO

def restaurar_abas():
    print(f"\n🔧 Iniciando Restauração de Abas em: {ARQUIVO_HISTORICO}")
    
    if not ARQUIVO_HISTORICO.exists():
        print("❌ Arquivo Excel não encontrado!")
        return

    try:
        wb = load_workbook(ARQUIVO_HISTORICO)
    except Exception as e:
        print(f"❌ Erro ao abrir arquivo: {e}")
        return

    abas_existentes = wb.sheetnames
    alteracoes = False
    
    # 1. RANKING INDICADORES
    if 'RANKING INDICADORES' not in abas_existentes:
        print("   ⚠️  Aba 'RANKING INDICADORES' ausente. Criando...")
        
        # Lista padrão de indicadores
        indicadores = [
            'Quadrantes', 'Div9', 'Fibonacci', 'Div6', 'Mult5', 'Div3', 'Gap', 'Primos', 
            'Simetria', 'ParImpar', 'Amplitude', 'Soma',
            'RaizDigital', 'VariacaoSoma', 'Conjugacao', 'RepeticaoAnterior', 'FrequenciaMensal',
            'Sequencias', 'DistanciaMedia', 'NumerosExtremos', 'PadraoDezena', 'CicloAparicao',
            'TendenciaQuadrantes', 'CiclosSemanais', 'AcumulacaoConsecutiva', 'JanelaDeslizante',
            'MatrizPosicional', 'ClusterEspacial', 'SimetriaCentral',
            'FrequenciaRelativa', 'DesvioFrequencia', 'EntrópiaDistribuicao', 'CorrelacaoTemporal',
            'SomaDigitos', 'PadraoModular',
            'ScoreAnomalia', 'ProbabilidadeCondicional', 'ImportanciaFeature',
            'PadroesSubconjuntos', 'MicroTendencias', 'AnaliseContextual', 'Embedding'
        ]
        
        df_ranking = pd.DataFrame({
            'Indicador': indicadores,
            'Peso_Atual': [50.0] * len(indicadores),
            'Descrição': ['-'] * len(indicadores),
            'Categoria': ['Geral'] * len(indicadores)
        })
        
        # Salvar (usando append/writer helper seria melhor, mas aqui faremos direto via Pandas e depois reabrimos)
        # Nota: Pandas writer com openpyxl mode='a' é o ideal
        ws = wb.create_sheet('RANKING INDICADORES')
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r in dataframe_to_rows(df_ranking, index=False, header=True):
            ws.append(r)
            
        alteracoes = True
        print("   ✅ Aba 'RANKING INDICADORES' criada com 42 indicadores padrão.")
    else:
        print("   ✅ Aba 'RANKING INDICADORES' já existe.")

    # 2. HISTÓRICO_PESOS
    if 'HISTÓRICO_PESOS' not in abas_existentes:
        print("   ⚠️  Aba 'HISTÓRICO_PESOS' ausente. Criando...")
        ws = wb.create_sheet('HISTÓRICO_PESOS')
        ws.append(['Data', 'Performance_Ciclo']) # Cabeçalho Mínimo
        alteracoes = True
        print("   ✅ Aba 'HISTÓRICO_PESOS' criada.")
    else:
        print("   ✅ Aba 'HISTÓRICO_PESOS' já existe.")

    # 3. ANÁLISE IA
    if 'ANÁLISE IA' not in abas_existentes:
        print("   ⚠️  Aba 'ANÁLISE IA' ausente. Criando...")
        ws = wb.create_sheet('ANÁLISE IA')
        ws.append(['Data', 'Tipo', 'Indicador', 'Conteudo', 'Performance_Ciclo'])
        alteracoes = True
        print("   ✅ Aba 'ANÁLISE IA' criada.")
    else:
        print("   ✅ Aba 'ANÁLISE IA' já existe.")

    if alteracoes:
        try:
            wb.save(ARQUIVO_HISTORICO)
            print("\n💾 Alterações salvas com sucesso!")
        except PermissionError:
            print("\n❌ ERRO DE PERMISSÃO: Feche o arquivo Excel e tente novamente.")
    else:
        print("\n✨ Nenhuma alteração necessária.")

if __name__ == "__main__":
    restaurar_abas()
